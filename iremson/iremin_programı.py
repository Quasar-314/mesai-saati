from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QCheckBox, QPushButton, QLabel, 
    QLineEdit, QTableWidget, QTableWidgetItem, QMessageBox, QComboBox, 
    QDateEdit, QApplication, QTabWidget, QDialog, QHBoxLayout, QGroupBox, 
    QFormLayout, QDialogButtonBox, QHeaderView, QCalendarWidget, QTextEdit
)

from PyQt6.QtCore import Qt, QDateTime, pyqtSignal, QSize, QDate
from PyQt6.QtGui import QColor, QIcon

from datetime import datetime
import sqlite3
import sys
import os
import pandas as pd

from docs.StyleEditor import StyleEditor
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from cryptography.fernet import Fernet




# Stil tanımlamaları - import satırlarından sonra ekleyin
MAIN_STYLE = """

            QMainWindow, QWidget {
                background-color: #91B5D5;
                color: #000000;
                font-family: 'Segoe UI';
                font-size: 12px;
            }

            QTabWidget::pane {
                border: 2px solid #CC3300;
                border-radius: 6px;
                background-color: white;
                padding: 5px;
            }

            QTabBar::tab {
                background-color: #E6E6E6;
                color: #000000;
                padding: 12px 25px;
                margin: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-size: 13px;
                font-weight: bold;
            }

            QTabBar::tab:selected {
                background-color: #CC3300;
                color: white;
            }

            QTabBar::tab:hover:!selected {
                background-color: #CCCCCC;
            }

            QTableWidget {
                background-color: white;
                alternate-background-color: #F5F5F5;
                border: 1px solid #CCCCCC;
                border-radius: 6px;
                gridline-color: #E0E0E0;
            }

            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #E0E0E0;
            }

            QTableWidget::item:selected {
                background-color: #CC3300;
                color: white;
            }

            QHeaderView::section {
                background-color: #4A4A4A;
                color: white;
                padding: 12px;
                font-weight: bold;
                border: none;
                border-right: 1px solid #666666;
            }

            QPushButton {
                background-color: #CC3300;
                color: #FFFFFF;
                border: none;
                border-radius: 4px;
                padding: 8px;
                font-weight: bold;
                min-width: 100px;
            }

            QPushButton:hover {
                background-color: #E63900;
            }

            QLineEdit, QDateEdit, QComboBox {
                padding: 8px;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                background-color: white;
            }
            """

LOGIN_STYLE = """
QDialog {
        background-color: white;
        border: 2px solid #ccc;
        border-radius: 10px;
}

QLineEdit {
    padding: 10px;
    border: 1px solid #cccccc;
    border-radius: 4px;
    font-size: 14px;
}

QPushButton {
    background-color: #CC3300;
    color: white;
    border: none;
    padding: 12px 24px;
    border-radius: 4px;
    font-weight: bold;
}
#container {
        background: white;
        border: 1px solid #ddd;
        border-radius: 25px;
    }
"""




def ensure_docs_structure():
    """Gerekli klasör yapısını oluşturur
    Bu fonksiyon, programın ihtiyaç duyduğu tüm klasörlerin varlığını kontrol eder
    ve eksik olanları oluşturur"""
    
    # Ana docs klasörünü oluştur
    if not os.path.exists('docs'):
        os.makedirs('docs')
    
    # Çalışan dosyaları için alt klasör oluştur
    if not os.path.exists('docs/employee_files'):
        os.makedirs('docs/employee_files')
    
    # İkonlar için alt klasör oluştur
    if not os.path.exists('docs/icon'):
        os.makedirs('docs/icon')


# Import satırlarının ve stil tanımlamalarının altına ekleyin
def setup_modern_table(table_widget):
    """Modern tablo ayarlarını yapılandır"""
    # Temel tablo ayarları
    table_widget.setAlternatingRowColors(True)
    table_widget.setShowGrid(True)
    table_widget.setGridStyle(Qt.PenStyle.SolidLine)
    table_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
    table_widget.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
    
    # Yatay başlık ayarları
    horizontal_header = table_widget.horizontalHeader()
    horizontal_header.setDefaultSectionSize(150)
    horizontal_header.setStretchLastSection(True)
    horizontal_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
    
    # Dikey başlık ayarları
    vertical_header = table_widget.verticalHeader()
    vertical_header.setDefaultSectionSize(40)
    vertical_header.setVisible(False)  # Satır numaralarını gizle
    
    # Köşe butonunu özelleştir
    table_widget.setCornerButtonEnabled(False)
    
    # Font ayarları
    font = table_widget.font()
    font.setPointSize(10)
    table_widget.setFont(font)






def create_tables(self):
    cursor = self.conn.cursor()
    
    # Simplified Users table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        department TEXT NOT NULL,  # 'Admin', 'Yönetici', or 'Muhasebe'
        can_view_employees BOOLEAN NOT NULL DEFAULT 0,
        can_manage_time_records BOOLEAN NOT NULL DEFAULT 0,
        can_view_employee_reports BOOLEAN NOT NULL DEFAULT 0,
        can_view_company_reports BOOLEAN NOT NULL DEFAULT 0,
        can_manage_users BOOLEAN NOT NULL DEFAULT 0
    )
    ''')

    cursor.execute('SELECT COUNT(*) FROM users')
    user_count = cursor.fetchone()[0]
    
    if user_count == 0:
        cursor.execute('''
        INSERT INTO users (
            username, password, department,
            can_view_employees, can_manage_time_records,
            can_view_employee_reports, can_view_company_reports,
            can_manage_users
        )
        VALUES (?, ?, 'Admin', 1, 1, 1, 1, 1)
        ''', ('admin', 'admin123'))

    self.conn.commit()  # Değişiklikleri kaydet

def create_label(text, is_header=False, is_form=False):
    """Özelleştirilmiş QLabel oluşturur"""
    label = QLabel(text)
    
    if is_header:
        label.setObjectName("headerLabel")
    elif is_form:
        label.setObjectName("formLabel")
    
    # Font ayarları
    font = label.font()
    font.setPointSize(12 if is_form else (14 if is_header else 13))
    font.setBold(True)
    label.setFont(font)
    
    return label


class LoginDialog(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.user_permissions = None
        self.username = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Giriş')
        self.setFixedSize(450, 600)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
    
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
    
        # Ana container
        container = QWidget()
        container.setObjectName("container")
        container_layout = QVBoxLayout()
        container_layout.setSpacing(20)  # Azalttık
        container_layout.setContentsMargins(40, 20, 40, 40)  # Üst margin'i azalttık
    
        # Çıkış butonu - En üstte
        close_button = QPushButton()
        close_button.setObjectName("closeButton")
        close_button.clicked.connect(self.reject)
        close_button.setCursor(Qt.CursorShape.PointingHandCursor)
        close_button.setFixedSize(30, 30)
        
        # Logo/İkon alanı
        icon_label = QLabel()
        icon_label.setFixedSize(100, 100)
        icon_label.setPixmap(QIcon("docs/icon/irem2.jpg").pixmap(100, 100))
        
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
    
        # Başlık - İki satıra böl
        title_container = QWidget()
        title_layout = QVBoxLayout()
        title_layout.setSpacing(5)  # Satırlar arası boşluk ekledik
        title_layout.setContentsMargins(0, 0, 0, 0)
    
        # Başlık etiketleri
        title_label1 = QLabel("İREMİN PROGRAMINA")
        title_label1.setObjectName("titleLabel")
        title_label1.setAlignment(Qt.AlignmentFlag.AlignCenter)
    
        title_label2 = QLabel("HOŞ GELDİNİZ")
        title_label2.setObjectName("titleLabel")
        title_label2.setAlignment(Qt.AlignmentFlag.AlignCenter)
    
        title_layout.addWidget(title_label1)
        title_layout.addWidget(title_label2)
        title_container.setLayout(title_layout)
    
        # Form container - Daha kompakt
        form_container = QWidget()
        form_container.setObjectName("formContainer")
        form_layout = QVBoxLayout()
        form_layout.setSpacing(20)  # Azalttık
        form_layout.setContentsMargins(30, 30, 30, 30)  # Azalttık
    
        # Giriş alanları
        self.username_input = QLineEdit()
        self.username_input.setObjectName("inputField")
        self.username_input.setPlaceholderText("Kullanıcı Adı")
        self.username_input.setMinimumHeight(55)
    
        self.password_input = QLineEdit()
        self.password_input.setObjectName("inputField")
        self.password_input.setPlaceholderText("Şifre")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setMinimumHeight(55)
    
        # Giriş butonu
        login_button = QPushButton("GİRİŞ YAP")
        login_button.setObjectName("loginButton")
        login_button.setIcon(QIcon("docs/icon/login.png"))  # İkon eklendi
        login_button.setIconSize(QSize(40, 40))  # İkon boyutu 32x32 olarak ayarlandı
        login_button.clicked.connect(self.check_login)
        login_button.setCursor(Qt.CursorShape.PointingHandCursor)
        login_button.setMinimumHeight(55)
    
        # Form elemanlarını ekle
        form_layout.addWidget(self.username_input)
        form_layout.addWidget(self.password_input)
        form_layout.addWidget(login_button)
        form_container.setLayout(form_layout)
    
        # Ana container'a elemanları ekle
        container_layout.addWidget(close_button, alignment=Qt.AlignmentFlag.AlignRight)
        container_layout.addWidget(icon_label, alignment=Qt.AlignmentFlag.AlignCenter)
        container_layout.addWidget(title_container)
        container_layout.addWidget(form_container)
        container.setLayout(container_layout)
    
        main_layout.addWidget(container)
        self.setLayout(main_layout)
    
        # Enter tuşu desteği
        self.username_input.returnPressed.connect(self.check_login)
        self.password_input.returnPressed.connect(self.check_login)
    
        # Stil tanımlamaları - Başlık stilini güncelledik
        # LoginDialog sınıfı içinde init_ui metodunda, setStyleSheet kısmını güncelle
        self.setStyleSheet("""
            QDialog {
                background: transparent;
            }
            
            #container {
                background: white;
                border-radius: 25px;
            }
            
            #titleLabel {
                color: #2C3E50;
                font-size: 24px;
                font-weight: bold;
                letter-spacing: 1px;
                margin: 5px 0;
            }
            
            #formContainer {
                background: #F8F9FA;
                border-radius: 20px;
                padding: 20px;
            }
            
            #inputField {
                border: 2px solid #E0E5EC;
                border-radius: 12px;
                padding: 15px 20px;
                font-size: 16px;
                color: #2C3E50;
                background: white;
                font-weight: 500;
            }
            
            #inputField:focus {
                border-color: #CC3300;
                outline: none;
                background: #FFFFFF;
            }
            
            #inputField::placeholder {
                color: #95A5A6;
                font-weight: 400;
            }
            
            #loginButton {
                background: #CC3300;
                border: none;
                border-radius: 12px;
                color: white;
                padding: 15px;
                font-size: 16px;
                font-weight: bold;
                letter-spacing: 1px;
                margin-top: 10px;
                text-transform: uppercase;
            }
            
            #loginButton:hover {
                background: #E63900;
            }
            
            #loginButton:pressed {
                background: #CC3300;
            }
            
            #closeButton {
                background: transparent;
                border: none;
                color: #95A5A6;
                font-size: 20px;
                font-weight: bold;
            }
            
            #closeButton:hover {
                color: #CC3300;
            }
            
            #closeButton::after {
                content: "×";
                position: absolute;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%);
                font-size: 28px;
            }
        """)

    def mousePressEvent(self, event):
        """Pencereyi sürüklemek için başlangıç pozisyonunu kaydet"""
        if event.button() == Qt.MouseButton.LeftButton:
            self.drag_position = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        """Pencereyi sürükle"""
        if event.buttons() & Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self.drag_position)
            event.accept()

    def check_login(self):
        """Kullanıcı girişini kontrol et"""
        username = self.username_input.text()
        password = self.password_input.text()
    
        if not username or not password:
            msg = QMessageBox()
            msg.setWindowTitle('Hata')
            msg.setText('Lütfen tüm alanları doldurun!')
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: white;
                    color: black;
                }
                QLabel {
                    color: black;
                }
                QPushButton {
                    background-color: #CC3300;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 12px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #E63900;
                }
            """)
            msg.exec()
            return
    
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
    
        if not user:
            msg = QMessageBox()
            msg.setWindowTitle('Help İrem')
            msg.setText('Hatalı Kullanıcı İremden kullanıcı talep ediniz')
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: white;
                    color: black;
                }
                QLabel {
                    color: black;
                }
                QPushButton {
                    background-color: #CC3300;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 12px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #E63900;
                }
            """)
            msg.exec()
            self.username_input.clear()
            self.password_input.clear()
            self.username_input.setFocus()
            return
    
        if user[2] != password:
            msg = QMessageBox()
            msg.setWindowTitle('İrem Yardım')
            msg.setText('Şifre yanlış! İremden doğru şifreyi alınız')
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: white;
                    color: black;
                }
                QLabel {
                    color: black;
                }
                QPushButton {
                    background-color: #CC3300;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 12px;
                    font-weight: bold;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #E63900;
                }
            """)
            msg.exec()
            self.password_input.clear()
            self.password_input.setFocus()
            return
    
        # Başarılı giriş
        department = user[3]
        self.user_permissions = {
            'department': department,
            'can_view_employees': user[4],
            'can_manage_time_records': user[5],
            'can_view_employee_reports': user[6],
            'can_view_company_reports': user[7],
            'can_manage_users': user[8]
        }
        self.username = username
        self.accept()



class UserManagementTab(QWidget):
    user_added = pyqtSignal()
    user_updated = pyqtSignal()
    user_deleted = pyqtSignal()
    def __init__(self, db, main_window=None):
        super().__init__()
        self.db = db
        self.main_window = main_window
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        # Ana container widget
        container = QWidget()
        container_layout = QHBoxLayout()
        container_layout.setSpacing(20)

        # Sol panel - Form alanları
        left_panel = QWidget()
        left_layout = QVBoxLayout()
        left_layout.setSpacing(10)

        # Form grubu
        form_group = QGroupBox("Yeni Kullanıcı Ekle")
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText('Kullanıcı Adı')
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText('Şifre')
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)

        # Departman seçimi
        self.department_combo = QComboBox()
        self.department_combo.addItems(['Muhasebe', 'Yönetici'])

        form_layout.addRow('Kullanıcı Adı:', self.username_input)
        form_layout.addRow('Şifre:', self.password_input)
        form_layout.addRow('Departman:', self.department_combo)

       


        # Butonlar için vertical layout
        button_layout = QVBoxLayout()
        button_layout.setSpacing(10)  # Butonlar arası boşluk
        
        # Kullanıcı Ekle butonu
        add_button = QPushButton('Kullanıcı Ekle')
        add_button.setIcon(QIcon("docs/icon/adduser.png"))
        add_button.setIconSize(QSize(20, 20))
        add_button.clicked.connect(self.add_user)
        add_button.setMinimumWidth(250)  # Genişliği artırıldı
        add_button.setMinimumHeight(40)  # Yüksekliği artırıldı
        
        # Kullanıcı Adı ve Şifre Değiştir butonu
        change_password_button = QPushButton('Kullanıcı Adı ve Şifre Değiştir')
        change_password_button.setIcon(QIcon("docs/icon/passwordedit.png"))
        change_password_button.setIconSize(QSize(20, 20))
        change_password_button.clicked.connect(self.change_password)
        change_password_button.setMinimumWidth(250)  # Genişliği artırıldı
        change_password_button.setMinimumHeight(40)  # Yüksekliği artırıldı
        
        # Stil editörü butonu
        self.style_editor_button = QPushButton('Stil Editörü')
        self.style_editor_button.setIcon(QIcon("docs/icon/art.png"))
        self.style_editor_button.setIconSize(QSize(20, 20))
        self.style_editor_button.clicked.connect(self.open_style_editor)
        self.style_editor_button.setMinimumWidth(250)
        self.style_editor_button.setMinimumHeight(40)
        self.style_editor_button.setStyleSheet("""
            QPushButton {
                background-color: #CC3300;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #E63900;
            }
            """)
        
        # Butonları layout'a ekle
        button_layout.addWidget(add_button)
        button_layout.addWidget(change_password_button)
        button_layout.addWidget(self.style_editor_button)
        button_layout.addStretch()  # Alt kısımda boşluk bırakır







        # Admin kontrolü
        if self.main_window and self.main_window.user_permissions:
            is_admin = self.main_window.user_permissions.get('department') == 'Admin'
            self.style_editor_button.setVisible(is_admin)

        button_layout.addWidget(add_button)
        button_layout.addWidget(change_password_button)
        button_layout.addWidget(self.style_editor_button)
        button_layout.addStretch()

        # Form grubuna layout'ları ekle
        form_group.setLayout(form_layout)
        left_layout.addWidget(form_group)
        left_layout.addLayout(button_layout)
        left_layout.addStretch()
        left_panel.setLayout(left_layout)

        # Sağ panel - Tablo
        right_panel = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setSpacing(10)

        # Tablo başlığı ve tablo
        table_group = QGroupBox("Kullanıcı Listesi")
        table_layout = QVBoxLayout()

        self.user_table = QTableWidget()
        self.user_table.setColumnCount(3)
        self.user_table.setHorizontalHeaderLabels(['ID', 'Kullanıcı Adı', 'Departman'])
        self.user_table.itemDoubleClicked.connect(self.edit_user)
        setup_modern_table(self.user_table)  # Modern tablo stilini uygula

        # Silme butonu
        delete_button = QPushButton('Seçili Kullanıcıyı Sil')
        delete_button.setIcon(QIcon("docs/icon/deletuser.png"))
        delete_button.setIconSize(QSize(20, 20))        
        delete_button.clicked.connect(self.delete_user)
        delete_button.setMinimumWidth(150)

        table_layout.addWidget(self.user_table)
        table_layout.addWidget(delete_button)
        table_group.setLayout(table_layout)

        right_layout.addWidget(table_group)
        right_panel.setLayout(right_layout)

        # Container'a panelleri ekle
        container_layout.addWidget(left_panel, 1)  # 1 birim genişlik
        container_layout.addWidget(right_panel, 2)  # 2 birim genişlik
        container.setLayout(container_layout)

        # Ana layout'a container'ı ekle
        layout.addWidget(container)
        self.setLayout(layout)

        # Kullanıcıları yükle
        self.load_users()

    def add_user(self):
        try:
            username = self.username_input.text()
            password = self.password_input.text()
            department = self.department_combo.currentText()
    
            if not all([username, password]):
                QMessageBox.warning(self, 'Hata', 'Kullanıcı adı ve şifre gerekli!')
                return
    
            # Departmana göre yetkileri belirle
            if department == 'Yönetici':
                permissions = {
                    'can_view_employees': True,
                    'can_manage_time_records': True,
                    'can_view_employee_reports': True,
                    'can_view_company_reports': True,
                    'can_manage_users': False
                }
            else:  # Muhasebe
                permissions = {
                    'can_view_employees': False,
                    'can_manage_time_records': True,
                    'can_view_employee_reports': False,
                    'can_view_company_reports': False,
                    'can_manage_users': False
                }
    
            cursor = self.db.conn.cursor()
            cursor.execute('''
                INSERT INTO users (
                    username, password, department,
                    can_view_employees, can_manage_time_records,
                    can_view_employee_reports, can_view_company_reports,
                    can_manage_users
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                username, password, department,
                permissions['can_view_employees'],
                permissions['can_manage_time_records'],
                permissions['can_view_employee_reports'],
                permissions['can_view_company_reports'],
                permissions['can_manage_users']
            ))
            
            # Burayı ekleyin - Yeni kullanıcının ID'sini al
            user_id = cursor.lastrowid
            
            # Burayı ekleyin - Log kaydı ekle
            if self.main_window and self.main_window.current_username:
                self.db.log_activity(
                    self.main_window.current_username,
                    "USER_ADD",
                    f"Yeni kullanıcı eklendi: {username}",
                    "users",  # Bu 'Kullanıcı' olarak görünecek
                    user_id
                )
                
            self.db.conn.commit()
            self.clear_form()
            self.load_users()
            QMessageBox.information(self, 'Başarılı', 'Kullanıcı eklendi!')
            
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, 'Hata', 'Bu kullanıcı adı zaten kullanımda!')
        
        self.user_added.emit()
        
    def open_style_editor(self):
        """Stil editörünü aç"""
        if self.main_window:
            self.main_window.style_editor.show()
        else:
            style_editor = StyleEditor()
            style_editor.show()

    def edit_user(self, item):
        row = item.row()
        user_id = self.user_table.item(row, 0).text()
        
        dialog = QDialog(self)
        dialog.setWindowTitle('Kullanıcı Düzenle')
        layout = QVBoxLayout()
        
        # Form elemanları
        form_layout = QFormLayout()
        username_input = QLineEdit(self.user_table.item(row, 1).text())
        department_combo = QComboBox()
        department_combo.addItems(['Muhasebe', 'Yönetici'])
        department_combo.setCurrentText(self.user_table.item(row, 2).text())
        
        form_layout.addRow('Kullanıcı Adı:', username_input)
        form_layout.addRow('Departman:', department_combo)
        
        # Butonlar
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        
        layout.addLayout(form_layout)
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                new_username = username_input.text()
                new_department = department_combo.currentText()
                
                # Yeni departmana göre yetkileri belirle
                if new_department == 'Yönetici':
                    permissions = {
                        'can_view_employees': True,
                        'can_manage_time_records': True,
                        'can_view_employee_reports': True,
                        'can_view_company_reports': True,
                        'can_manage_users': False
                    }
                else:  # Muhasebe
                    permissions = {
                        'can_view_employees': False,
                        'can_manage_time_records': True,
                        'can_view_employee_reports': False,
                        'can_view_company_reports': False,
                        'can_manage_users': False
                    }

                cursor = self.db.conn.cursor()
                cursor.execute('''
                    UPDATE users 
                    SET username = ?, department = ?,
                        can_view_employees = ?, can_manage_time_records = ?,
                        can_view_employee_reports = ?, can_view_company_reports = ?,
                        can_manage_users = ?
                    WHERE id = ?
                ''', (
                    new_username, new_department,
                    permissions['can_view_employees'],
                    permissions['can_manage_time_records'],
                    permissions['can_view_employee_reports'],
                    permissions['can_view_company_reports'],
                    permissions['can_manage_users'],
                    user_id
                ))
                
                self.db.conn.commit()
                self.load_users()
                QMessageBox.information(self, 'Başarılı', 'Kullanıcı güncellendi!')
            
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, 'Hata', 'Bu kullanıcı adı zaten kullanımda!')
            except Exception as e:
                QMessageBox.warning(self, 'Hata', f'Güncelleme hatası: {str(e)}')

    def delete_user(self):
        selected_items = self.user_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'Uyarı', 'Lütfen silinecek bir kullanıcı seçin!')
            return

        row = selected_items[0].row()
        user_id = self.user_table.item(row, 0).text()
        username = self.user_table.item(row, 1).text()

        reply = QMessageBox.question(
            self, 'Onay',
            f'{username} kullanıcısını silmek istediğinizden emin misiniz?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            cursor = self.db.conn.cursor()
            cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
            self.db.conn.commit()
            self.load_users()
            QMessageBox.information(self, 'Başarılı', 'Kullanıcı silindi!')

        
        self.user_deleted.emit()

    def clear_form(self):
        self.username_input.clear()
        self.password_input.clear()
        self.department_combo.setCurrentIndex(0)

    def change_password(self):
        if self.main_window:
            dialog = ChangeUserDetailsDialog(self.db, self.main_window.current_username, self.main_window)
        else:
            dialog = ChangeUserDetailsDialog(self.db, 'admin')
        dialog.exec()

    def load_users(self):
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id, username, department FROM users WHERE department != "Admin"')
        users = cursor.fetchall()
    
        self.user_table.setRowCount(len(users))
        for row, user in enumerate(users):
            for col, value in enumerate(user):
                self.user_table.setItem(row, col, QTableWidgetItem(str(value)))



class ChangeUserDetailsDialog(QDialog):
    def __init__(self, db, username, parent=None):
        super().__init__(parent)
        self.db = db
        self.username = username
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Kullanıcı Bilgilerini Değiştir')
        layout = QVBoxLayout()

        # Mevcut şifre alanı
        self.current_password = QLineEdit()
        self.current_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.current_password.setPlaceholderText('Mevcut Şifre')

        # Yeni kullanıcı adı alanı
        self.new_username = QLineEdit()
        self.new_username.setPlaceholderText('Yeni Kullanıcı Adı (opsiyonel)')

        # Yeni şifre alanları
        self.new_password = QLineEdit()
        self.new_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.new_password.setPlaceholderText('Yeni Şifre (opsiyonel)')

        self.confirm_password = QLineEdit()
        self.confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.confirm_password.setPlaceholderText('Yeni Şifre (Tekrar)')

        # Buton düzeni
        button_layout = QHBoxLayout()
        
        # Kaydet butonu
        save_button = QPushButton('Değişiklikleri Kaydet')
        save_button.clicked.connect(self.change_user_details)
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 5px;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # İptal butonu
        cancel_button = QPushButton('İptal')
        cancel_button.clicked.connect(self.reject)
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 5px;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)

        # Layout'a widgetları ekle
        layout.addWidget(create_label('Mevcut Şifre:', is_form=True))
        layout.addWidget(self.current_password)
        layout.addWidget(create_label('Yeni Kullanıcı Adı:', is_form=True))
        layout.addWidget(self.new_username)
        layout.addWidget(create_label('Yeni Şifre:', is_form=True))
        layout.addWidget(self.new_password)
        layout.addWidget(create_label('Yeni Şifre (Tekrar):', is_form=True))
        layout.addWidget(self.confirm_password)
        layout.addLayout(button_layout)

        self.setLayout(layout)

    def change_user_details(self):
        current = self.current_password.text()
        new_username = self.new_username.text()
        new_pass = self.new_password.text()
        confirm = self.confirm_password.text()

        if not current:
            QMessageBox.warning(self, 'Hata', 'Mevcut şifre gerekli!')
            return

        if not new_username and not new_pass:
            QMessageBox.warning(self, 'Hata', 'En az bir değişiklik yapmalısınız!')
            return

        if new_pass and new_pass != confirm:
            QMessageBox.warning(self, 'Hata', 'Yeni şifreler eşleşmiyor!')
            return

        # Mevcut şifreyi ve kullanıcı bilgilerini kontrol et
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id, is_admin, role FROM users WHERE username = ? AND password = ?',
                      (self.username, current))
        
        user = cursor.fetchone()
        if not user:
            QMessageBox.warning(self, 'Hata', 'Mevcut şifre yanlış!')
            return

        try:
            # Kullanıcı adı değişikliğini kontrol et
            if new_username:
                cursor.execute('SELECT id FROM users WHERE username = ? AND username != ?',
                             (new_username, self.username))
                if cursor.fetchone():
                    QMessageBox.warning(self, 'Hata', 'Bu kullanıcı adı zaten kullanımda!')
                    return

            # Değişiklikleri uygula
            update_fields = []
            params = []
            
            if new_username:
                update_fields.append('username = ?')
                params.append(new_username)
            
            if new_pass:
                update_fields.append('password = ?')
                params.append(new_pass)
            
            # Admin bilgisini koru
            update_fields.append('is_admin = ?')
            params.append(user[1])
            
            # Rol bilgisini koru
            update_fields.append('role = ?')
            params.append(user[2])
            
            params.append(self.username)  # WHERE koşulu için
            
            query = f'''UPDATE users 
                       SET {', '.join(update_fields)}
                       WHERE username = ?'''
            
            cursor.execute(query, params)
            self.db.conn.commit()
            
            reply = QMessageBox.information(
                self, 
                'Başarılı', 
                'Kullanıcı bilgileri güncellendi! Değişikliklerin uygulanması için program yeniden başlatılacak.',
                QMessageBox.StandardButton.Ok
            )
            
            self.accept()
            
            # Programı yeniden başlat
            if isinstance(self.parent, MainWindow):
                self.parent.restart_application()
            
        except Exception as e:
            QMessageBox.critical(self, 'Hata', f'Güncelleme başarısız: {str(e)}')




class SecuritySystem:
    def __init__(self):
        self.key_file = "docs/secure_key.key"  # docs klasörüne taşındı
        self._create_or_load_key()

    def _create_or_load_key(self):
        """Güvenlik anahtarı oluştur veya yükle"""
        if not os.path.exists(self.key_file):
            # Yeni anahtar oluştur
            self.key = Fernet.generate_key()
            with open(self.key_file, 'wb') as key_file:
                key_file.write(self.key)
        else:
            # Var olan anahtarı yükle
            with open(self.key_file, 'rb') as key_file:
                self.key = key_file.read()
        
        self.fernet = Fernet(self.key)

    def encrypt_file(self, file_path):
        """Dosyayı şifrele"""
        try:
            # Dosyayı oku
            with open(file_path, 'rb') as file:
                file_data = file.read()
            
            # Şifrele
            encrypted_data = self.fernet.encrypt(file_data)
            
            # Şifrelenmiş dosyayı kaydet
            with open(file_path + '.encrypted', 'wb') as file:
                file.write(encrypted_data)
            
            # Orijinal dosyayı sil
            os.remove(file_path)
            return True
        except Exception as e:
            print(f"Şifreleme hatası: {str(e)}")
            return False

    def decrypt_file(self, encrypted_file_path):
        """Şifrelenmiş dosyayı çöz"""
        try:
            # Şifrelenmiş dosyayı oku
            with open(encrypted_file_path, 'rb') as file:
                encrypted_data = file.read()
            
            # Şifreyi çöz
            decrypted_data = self.fernet.decrypt(encrypted_data)
            
            # Orijinal dosyayı oluştur
            original_file = encrypted_file_path.replace('.encrypted', '')
            with open(original_file, 'wb') as file:
                file.write(decrypted_data)
            
            # Şifrelenmiş dosyayı sil
            os.remove(encrypted_file_path)
            return True
        except Exception as e:
            print(f"Şifre çözme hatası: {str(e)}")
            return False

    def secure_directory(self, directory_path):
        """Bir klasördeki tüm dosyaları şifrele"""
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                if not file.endswith('.encrypted') and not file.endswith('.key'):
                    file_path = os.path.join(root, file)
                    self.encrypt_file(file_path)

    def unsecure_directory(self, directory_path):
        """Bir klasördeki tüm şifrelenmiş dosyaların şifresini çöz"""
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                if file.endswith('.encrypted'):
                    file_path = os.path.join(root, file)
                    self.decrypt_file(file_path)

class Database:
    def __init__(self):
        ensure_docs_structure
        self.security = SecuritySystem()  # Güvenlik sistemi ekle
        self.conn = None
        self.connect_and_decrypt()  # Başlangıçta bağlantıyı ve şifre çözmeyi başlat
        self.create_tables()

    def connect_and_decrypt(self):
        """Dosyaların şifresini çöz ve veritabanına bağlan"""
        try:
            # Veritabanı şifresini çöz
            if os.path.exists('docs/employee_management.db.encrypted'):
                self.security.decrypt_file('docs/employee_management.db.encrypted')
            
            # Employee files klasörünün şifresini çöz
            if os.path.exists('docs/employee_files'):
                self.security.unsecure_directory('docs/employee_files')
            
            # Veritabanına bağlan
            self.conn = sqlite3.connect('docs/employee_management.db')
            return True
        except Exception as e:
            print(f"Bağlantı hatası: {str(e)}")
            return False

    def close(self):
        """Veritabanını kapat ve dosyaları şifrele"""
        if self.conn:
            self.conn.close()
            try:
                # employee_files klasörünü şifrele
                if os.path.exists('employee_files'):
                    self.security.secure_directory('employee_files')
                # Veritabanı dosyasını şifrele
                if os.path.exists('employee_management.db'):
                    self.security.encrypt_file('employee_management.db')
            except Exception as e:
                print(f"Şifreleme hatası: {str(e)}")

    def create_tables(self):
        cursor = self.conn.cursor()
        
        # Kullanıcılar tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            department TEXT NOT NULL,  
            can_view_employees BOOLEAN NOT NULL DEFAULT 0,
            can_manage_time_records BOOLEAN NOT NULL DEFAULT 0,
            can_view_employee_reports BOOLEAN NOT NULL DEFAULT 0,
            can_view_company_reports BOOLEAN NOT NULL DEFAULT 0,
            can_manage_users BOOLEAN NOT NULL DEFAULT 0
        )
        ''')

        # Admin kullanıcısını kontrol et ve ekle
        cursor.execute('SELECT COUNT(*) FROM users WHERE department = "Admin"')
        admin_count = cursor.fetchone()[0]
        
        if admin_count == 0:
            cursor.execute('''
            INSERT INTO users (
                username, password, department,
                can_view_employees, can_manage_time_records,
                can_view_employee_reports, can_view_company_reports,
                can_manage_users
            )
            VALUES (?, ?, 'Admin', 1, 1, 1, 1, 1)
            ''', ('admin', 'admin123'))

        # Çalışanlar tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            position TEXT NOT NULL,
            salary REAL NOT NULL,
            hire_date TEXT NOT NULL
        )
        ''')

        
        

        # Tatiller tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            description TEXT NOT NULL
        )
        ''')

        self.conn.commit()
        # Çalışanlar tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            position TEXT NOT NULL,
            salary REAL NOT NULL,
            hire_date TEXT NOT NULL
        )
        ''')

         # Mesai kayıtları tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS time_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            date TEXT NOT NULL,
            check_in TEXT NOT NULL,
            check_out TEXT NOT NULL,
            working_hours REAL NOT NULL,
            overtime_hours REAL NOT NULL,
            is_holiday BOOLEAN NOT NULL DEFAULT 0,
            is_weekend BOOLEAN NOT NULL DEFAULT 0,
            base_pay REAL NOT NULL DEFAULT 0,
            overtime_pay REAL NOT NULL DEFAULT 0,
            total_pay REAL NOT NULL DEFAULT 0,
            is_sick_leave BOOLEAN NOT NULL DEFAULT 0,
            is_paid_leave BOOLEAN NOT NULL DEFAULT 0,
            is_weekend_leave BOOLEAN NOT NULL DEFAULT 0,
            FOREIGN KEY (employee_id) REFERENCES employees (id)
        )
        ''')

        # Tatiller tablosu
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            description TEXT NOT NULL
        )
        ''')

        self.conn.commit()


        # Activity Logs tablosunu güncelle
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS activity_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            user_id INTEGER,
            username TEXT,
            action_type TEXT,
            action_description TEXT,
            related_table TEXT,
            related_id INTEGER,
            previous_state TEXT,  -- Yeni kolon
            new_state TEXT,      -- Yeni kolon
            change_reason TEXT,  -- Yeni kolon
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        ''')

        # Mevcut kayıtları yeni kolonlarla uyumlu hale getir
        try:
            cursor.execute("ALTER TABLE activity_logs ADD COLUMN previous_state TEXT")
            cursor.execute("ALTER TABLE activity_logs ADD COLUMN new_state TEXT")
            cursor.execute("ALTER TABLE activity_logs ADD COLUMN change_reason TEXT")
        except sqlite3.OperationalError:
            # Kolonlar zaten varsa hata almayı görmezden gel
            pass

        self.conn.commit()


    
    
    # Database sınıfı içinde update_time_records_table çağrısını düzelt
    def update_time_records_table(self):  # self parametresi ekle
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA table_info(time_records)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'is_weekend_leave' not in columns:
            cursor.execute('''
                ALTER TABLE time_records 
                ADD COLUMN is_weekend_leave BOOLEAN NOT NULL DEFAULT 0
            ''')
        self.conn.commit()
    
    def add_employee(self, name, surname, position, salary):
        salary_formatted = f"{salary:,.0f}"  # Virgüllü, ondalık basamaksız format
        cursor = self.conn.cursor()
        hire_date = datetime.now().strftime('%Y-%m-%d')
        cursor.execute('''
        INSERT INTO employees (name, surname, position, salary, hire_date)
        VALUES (?, ?, ?, ?, ?)
        ''', (name, surname, position, salary_formatted, hire_date))
        self.conn.commit()
        return cursor.lastrowid
    



    def get_all_employees(self):
        """Tüm çalışanları getirir"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM employees ORDER BY id')
        return cursor.fetchall()

    def get_employee(self, emp_id):
        """Belirli bir çalışanın bilgilerini getirir"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM employees WHERE id = ?', (emp_id,))
        return cursor.fetchone()

    def add_holiday(self, date, description):
        """Tatil günü ekler"""
        cursor = self.conn.cursor()
        cursor.execute('INSERT OR IGNORE INTO holidays (date, description) VALUES (?, ?)',
                      (date, description))
        self.conn.commit()

    def is_holiday(self, date):
        """Tarihin tatil günü olup olmadığını kontrol eder"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id FROM holidays WHERE date = ?', (date,))
        return cursor.fetchone() is not None

    def is_weekend(self, date):
        """Tarihin hafta sonu olup olmadığını kontrol eder"""
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        return date_obj.weekday() >= 5

    def calculate_working_hours(self, check_in, check_out, date):
        """Çalışma saatlerini hesaplar"""
        time_format = '%H:%M'
        check_in_time = datetime.strptime(check_in, time_format)
        check_out_time = datetime.strptime(check_out, time_format)
        
        if check_out_time < check_in_time:
            check_out_time = check_out_time.replace(day=check_out_time.day + 1)
        
        total_hours = (check_out_time - check_in_time).total_seconds() / 3600
        
        # Cumartesi kontrolü
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        is_saturday = date_obj.weekday() == 5
        
        if is_saturday:
            standard_hours = 5  # Cumartesi için standart saat
            break_time = 0     # Cumartesi mola yok
            working_hours = total_hours  # Mola çıkarılmıyor
            overtime = max(0, working_hours - standard_hours)  # 5 saat üzeri mesai
        else:
            standard_hours = 8  # Normal günler için standart saat
            break_time = 1.5   # Normal günler için mola süresi
            working_hours = total_hours - break_time
            overtime = max(0, working_hours - standard_hours)
        
        return working_hours, overtime


    def calculate_pay(self, employee_id, working_hours, overtime_hours, is_holiday, is_weekend):
        cursor = self.conn.cursor()
        cursor.execute('SELECT salary FROM employees WHERE id = ?', (employee_id,))
        monthly_salary = cursor.fetchone()[0]
        monthly_salary = float(monthly_salary.replace(',', ''))
    
        # Günlük ücret hesaplama
        daily_rate = monthly_salary / 30
        hourly_rate = daily_rate / 8
        overtime_rate = hourly_rate * 1.5
        holiday_rate = hourly_rate * 2
    
        if is_holiday:
            base_pay = working_hours * holiday_rate
            overtime_pay = 0
        else:
            if is_weekend:  # Cumartesi günü için
                standard_hours = 5
                base_pay = daily_rate  # Tam günlük ücret (5 saat çalışsa bile)
                if working_hours > standard_hours:
                    overtime_pay = overtime_hours * overtime_rate
                else:
                    overtime_pay = 0
            else:  # Normal günler için
                standard_hours = 8
                if working_hours < standard_hours:
                    base_pay = working_hours * hourly_rate  # Eksik mesai için orantılı ödeme
                    overtime_pay = 0
                else:
                    base_pay = daily_rate  # Tam günlük ücret
                    overtime_pay = overtime_hours * overtime_rate
                
        total_pay = base_pay + overtime_pay
        return base_pay, overtime_pay, total_pay
    def add_time_record(self, employee_id, date, check_in, check_out):
        try:
            # Çalışma saatlerini hesapla
            working_hours, overtime_hours = self.calculate_working_hours(check_in, check_out, date)

            # Tatil ve hafta sonu kontrolü
            is_holiday = self.is_holiday(date)
            is_weekend = self.is_weekend(date)

            # İzinli ve raporlu durumları kontrolü
            cursor = self.conn.cursor()
            is_sick_leave = False
            is_paid_leave = False

            if check_in == "09:00" and check_out == "18:00" and working_hours == 8:
                is_sick_leave = True
            elif check_in == "00:00" and check_out == "00:00" and working_hours == 0:
                is_paid_leave = True

            # Ücretleri hesapla
            base_pay, overtime_pay, total_pay = self.calculate_pay(
                employee_id, working_hours, overtime_hours, is_holiday, is_weekend
            )

            cursor.execute('''
            INSERT INTO time_records (
                employee_id, date, check_in, check_out, working_hours, 
                overtime_hours, is_holiday, is_weekend, base_pay, overtime_pay, total_pay,
                is_sick_leave, is_paid_leave
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (employee_id, date, check_in, check_out, working_hours, 
                  overtime_hours, is_holiday, is_weekend, base_pay, overtime_pay, total_pay,
                  is_sick_leave, is_paid_leave))

            record_id = cursor.lastrowid
            self.conn.commit()
            return record_id

        except Exception as e:
            print(f"Mesai kaydı eklenirken hata: {str(e)}")
            return None

    def get_monthly_records(self, employee_id, month, year):
        """Aylık mesai kayıtlarını getirir"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT *, is_sick_leave, is_paid_leave 
            FROM time_records 
            WHERE employee_id = ? 
            AND strftime('%m', date) = ? 
            AND strftime('%Y', date) = ?
            ORDER BY date, check_in
        ''', (employee_id, f"{month:02d}", str(year)))
        return cursor.fetchall()



    
    def update_time_record(self, record_id, check_in, check_out):
        """Mesai kaydını güncelle"""
        # Önce mevcut kaydı al
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT employee_id, date, is_holiday, is_weekend
            FROM time_records WHERE id = ?
        ''', (record_id,))
        record = cursor.fetchone()

        if not record:
            return False

        employee_id, date, is_holiday, is_weekend = record

        # Yeni çalışma saatlerini hesapla
        working_hours, overtime_hours = self.calculate_working_hours(check_in, check_out, date)

        # Yeni ücretleri hesapla
        base_pay, overtime_pay, total_pay = self.calculate_pay(
            employee_id, working_hours, overtime_hours, is_holiday, is_weekend
        )

        # Kaydı güncelle
        cursor.execute('''
            UPDATE time_records 
            SET check_in = ?, check_out = ?, working_hours = ?, 
                overtime_hours = ?, base_pay = ?, overtime_pay = ?, total_pay = ?
            WHERE id = ?
        ''', (check_in, check_out, working_hours, overtime_hours,
              base_pay, overtime_pay, total_pay, record_id))

        self.conn.commit()
        return True
    


    def get_employee_monthly_summary(self, employee_id, month, year):
        """Çalışanın aylık özet bilgilerini getir"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT 
                COUNT(*) as total_days,
                SUM(working_hours) as total_hours,
                SUM(overtime_hours) as total_overtime,
                SUM(base_pay) as total_base_pay,
                SUM(overtime_pay) as total_overtime_pay,
                SUM(total_pay) as total_pay
            FROM time_records 
            WHERE employee_id = ? 
            AND strftime('%m', date) = ? 
            AND strftime('%Y', date) = ?
        ''', (employee_id, f"{month:02d}", str(year)))
        return cursor.fetchone()
    

    def get_company_monthly_totals(self, month, year):
        """İşletmenin aylık toplam ödemelerini hesapla"""
        cursor = self.conn.cursor()

        # Tüm çalışanların maaş toplamı
        cursor.execute('SELECT SUM(salary) FROM employees')
        total_salaries = cursor.fetchone()[0] or 0

        # Aylık mesai ödemeleri toplamı
        cursor.execute('''
            SELECT 
                COUNT(DISTINCT employee_id) as total_employees,
                SUM(working_hours) as total_hours,
                SUM(overtime_hours) as total_overtime,
                SUM(base_pay) as total_base_pay,
                SUM(overtime_pay) as total_overtime_pay,
                SUM(total_pay) as total_pay
            FROM time_records 
            WHERE strftime('%m', date) = ? 
            AND strftime('%Y', date) = ?
        ''', (f"{month:02d}", str(year)))

        result = cursor.fetchone()
        if result:
            return {
                'total_employees': result[0] or 0,
                'total_hours': result[1] or 0,
                'total_overtime': result[2] or 0,
                'total_base_pay': result[3] or 0,
                'total_overtime_pay': result[4] or 0,
                'total_monthly_pay': result[5] or 0,
                'total_salaries': total_salaries,
                'grand_total': (result[5] or 0) + total_salaries
            }
        return None
    

    # Database sınıfı içinde log_activity metodunu güncelleyin
    def log_activity(self, username, action_type, description, related_table=None, related_id=None):
        try:
            cursor = self.conn.cursor()
            
            # Tablo isimlerini Türkçeleştir
            table_names = {
                'users': 'Kullanıcı',
                'time_records': 'Mesai Saati',
                'employees': 'Çalışan'
            }
            
            # Action type'ları Türkçeleştir
            action_types = {
                'USER_ADD': 'Kullanıcı Ekleme',
                'TIME_RECORD_ADD': 'Mesai Kaydı Ekleme',
                'EMPLOYEE_ADD': 'Çalışan Ekleme',
                'USER_UPDATE': 'Kullanıcı Güncelleme',
                'EMPLOYEE_UPDATE': 'Çalışan Güncelleme',
                'TIME_RECORD_UPDATE': 'Mesai Kaydı Güncelleme',
                'USER_DELETE': 'Kullanıcı Silme',
                'EMPLOYEE_DELETE': 'Çalışan Silme'
            }
    
            # Türkçe karşılıkları al
            turkish_table = table_names.get(related_table, related_table)
            turkish_action = action_types.get(action_type, action_type)
    
            # Kullanıcı ID'sini al
            cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
            user = cursor.fetchone()
            user_id = user[0] if user else None
    
            # Log kaydını oluştur
            formatted_description = f"{description} - {turkish_table} bölümünde"
    
            cursor.execute('''
                INSERT INTO activity_logs (
                    user_id, username, action_type, action_description, 
                    related_table, related_id
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (user_id, username, turkish_action, formatted_description, turkish_table, related_id))
    
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Log kayıt hatası: {str(e)}")
            return False






class EmployeeFileSystem:   
    def __init__(self, base_dir='docs/employee_files', db=None):  # Yolu güncellendi
        self.base_dir = base_dir
        self.db = db  # veritabanı bağlantısını sakla
        self._ensure_base_directory()
        
        # Renk kodları tanımla
        self.colors = {
            'header': '366092',  # Koyu mavi
            'holiday': 'FFD700',  # Altın sarısı (Tatil/Bayram)
            'overtime': '90EE90',  # Açık yeşil (Fazla mesai)
            'missing': 'FFB6C1',  # Açık kırmızı (Eksik mesai)
            'normal': 'FFFFFF'    # Beyaz (Normal mesai)
        }

    def _ensure_base_directory(self):
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)

    def _format_excel_worksheet(self, worksheet, record_types=None):
        """Excel çalışma sayfasını formatla ve satırları renklendir"""
        # Başlık stili
        header_fill = PatternFill(start_color=self.colors['header'], 
                                end_color=self.colors['header'], 
                                fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sütun genişliklerini ayarla
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Başlıkları formatla
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Satırları renklendir ve formatla
        if record_types:
            for row_idx, record_type in record_types.items():
                row = worksheet[row_idx + 2]  # +2 çünkü başlık satırı var
                color = self.colors.get(record_type, self.colors['normal'])
                row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                for cell in row:
                    cell.fill = row_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # Record types yoksa normal formatlama yap
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def update_employee_timesheet(self, employee_id, name, surname, record_data):
        """Mesai kaydını Excel'de güncelle"""
        try:
            # Klasör yolunu kontrol et ve gerekirse oluştur
            folder_path = os.path.join(self.base_dir, f"{name}_{surname}_{employee_id}")
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
            
            # Excel dosya yolunu belirle
            record_date = datetime.strptime(record_data['date'], '%Y-%m-%d')
            month = record_date.month
            year = record_date.year
            month_year = f"{year}_{month:02d}"
            month_name = record_date.strftime('%B')
            excel_path = os.path.join(folder_path, f"{month_year}_{month_name}.xlsx")

            # Excel dosyası yoksa oluştur
            if not os.path.exists(excel_path):
                # Mesai kayıtları için DataFrame oluştur
                columns = [
                    'Çalışan Adı', 'Çalışan Soyadı', 'Tarih', 'Giriş Saati', 'Çıkış Saati',
                    'Çalışma Süresi (Saat)', 'Normal Mesai', 'Fazla Mesai', 'Eksik Mesai',
                    'Mesai Tipi', 'Normal Mesai Ücreti', 'Fazla Mesai Ücreti',
                    'Eksik Mesai Kesintisi', 'Günlük Net Ücret'
                ]
                df = pd.DataFrame(columns=columns)
                df.to_excel(excel_path, sheet_name='Mesai Kayıtları', index=False)

            # Mevcut kayıtları oku
            existing_records = pd.read_excel(excel_path)

            # Yeni kaydı oluştur
            new_record = {
                'Çalışan Adı': name,
                'Çalışan Soyadı': surname,
                'Tarih': record_data['date'],
                'Giriş Saati': record_data['check_in'],
                'Çıkış Saati': record_data['check_out'],
                'Çalışma Süresi (Saat)': f"{record_data['working_hours']:.0f}",
                'Normal Mesai': f"{record_data['regular_hours']:.0f}",
                'Fazla Mesai': f"{record_data['overtime_hours']:.0f}",
                'Eksik Mesai': f"{record_data['missing_hours']:.0f}",
                'Mesai Tipi': record_data['work_type'],
                'Normal Mesai Ücreti': f"{record_data['regular_pay']:,.2f} TL",
                'Fazla Mesai Ücreti': f"{record_data['overtime_pay']:,.2f} TL",
                'Eksik Mesai Kesintisi': f"{record_data['missing_pay']:,.2f} TL",
                'Günlük Net Ücret': f"{record_data['total_pay']:,.2f} TL"
            }

            # Tarihe göre mevcut kaydı bul ve güncelle
            date_mask = existing_records['Tarih'] == record_data['date']
            if date_mask.any():
                # Mevcut kaydı güncelle
                for col, value in new_record.items():
                    existing_records.loc[date_mask, col] = value
            else:
                # Tarih bulunamazsa yeni kayıt ekle
                existing_records = pd.concat([existing_records, pd.DataFrame([new_record])], ignore_index=True)
            
            # Tarihe göre sırala
            existing_records = existing_records.sort_values('Tarih')

            # Kaydet ve formatla
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                existing_records.to_excel(writer, sheet_name='Mesai Kayıtları', index=False)
                workbook = writer.book
                worksheet = workbook['Mesai Kayıtları']

                # Renklendirme için satır tiplerini belirle
                record_types = {}
                for idx in range(len(existing_records)):
                    row = existing_records.iloc[idx]
                    if row['Mesai Tipi'] in ['Tatil', 'Bayram']:
                        record_types[idx] = 'holiday'
                    elif float(str(row['Fazla Mesai']).replace(' ', '').replace('Saat', '')) > 0:
                        record_types[idx] = 'overtime'
                    elif float(str(row['Eksik Mesai']).replace(' ', '').replace('Saat', '')) > 0:
                        record_types[idx] = 'missing'
                    else:
                        record_types[idx] = 'normal'

                self._format_excel_worksheet(worksheet, record_types)

            return True

        except Exception as e:
            print(f"Excel güncelleme hatası: {str(e)}")
            return False
        


    def create_employee_folder(self, name, surname, employee_id):
        """Çalışan için klasör oluştur"""
        folder_name = f"{name}_{surname}_{employee_id}"
        folder_path = os.path.join(self.base_dir, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        return folder_path
    

    def create_monthly_excel(self, name, surname, employee_id, position, salary, month=None, year=None):
        """Belirli bir ay için Excel dosyası oluştur"""
        if month is None or year is None:
            current_date = datetime.now()
            month = current_date.month
            year = current_date.year
            
        folder_path = os.path.join(self.base_dir, f"{name}_{surname}_{employee_id}")
        month_year = f"{year}_{month:02d}"
        month_date = datetime(year, month, 1)
        month_name = month_date.strftime('%B')
        excel_path = os.path.join(folder_path, f"{month_year}_{month_name}.xlsx")
    
        if not os.path.exists(excel_path):
            # Mesai kayıtları için boş DataFrame oluştur
            timesheet_columns = [
                'Çalışan Adı', 'Çalışan Soyadı', 'Tarih', 'Giriş Saati', 'Çıkış Saati',
                'Çalışma Süresi (Saat)', 'Normal Mesai', 'Fazla Mesai', 'Eksik Mesai',
                'Mesai Tipi', 'Normal Mesai Ücreti', 'Fazla Mesai Ücreti',
                'Eksik Mesai Kesintisi', 'Günlük Net Ücret'
            ]
            timesheet_df = pd.DataFrame(columns=timesheet_columns)
            
            # Excel dosyasını oluştur ve formatla
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                timesheet_df.to_excel(writer, sheet_name='Mesai Kayıtları', index=False)
                
                workbook = writer.book
                worksheet = workbook['Mesai Kayıtları']
                self._format_excel_worksheet(worksheet)
    
        return excel_path
    


class EmployeeTab(QWidget):
    employee_added = pyqtSignal()
    employee_updated = pyqtSignal()
    employee_deleted = pyqtSignal()

    def __init__(self, db, main_window=None):  # main_window parametresini ekle
        super().__init__()
        self.db = db
        self.main_window = main_window  # main_window'u kaydet
        self.file_system = EmployeeFileSystem(db=db)
        self.init_ui()

    

    def init_ui(self):
        layout = QVBoxLayout()
        
        form_widget = QWidget()
        form_layout = QVBoxLayout()
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText('Ad')
        self.surname_input = QLineEdit()
        self.surname_input.setPlaceholderText('Soyad')
        self.position_input = QLineEdit()
        self.position_input.setPlaceholderText('Pozisyon')
        
        self.salary_input = QLineEdit()
        self.salary_input.setPlaceholderText('Maaş')
        self.salary_input.textChanged.connect(self.format_salary_input)

        add_button = QPushButton('Çalışan Ekle')
        add_button.clicked.connect(self.add_employee)
        add_button.setIcon(QIcon("docs/icon/employeesadd.png"))
        add_button.setIconSize(QSize(20, 20))
        
        # İkon boyutunu ayarla
        icon_size = QSize(24, 24)  # İkon boyutunu 24x24 piksel yap
        add_button.setIconSize(icon_size)
        
        # Buton padding'ini ayarla (iç boşluk)
        add_button.setStyleSheet("""
            QPushButton {
                padding: 8px 16px 8px 8px;  /* üst sağ alt sol */
                min-height: 40px;
                min-width: 120px;
            }
        """)

        form_layout.addWidget(create_label('Yeni Çalışan Ekle', is_form=True))
        form_layout.addWidget(self.name_input)
        form_layout.addWidget(self.surname_input)
        form_layout.addWidget(self.position_input)
        form_layout.addWidget(self.salary_input)
        form_layout.addWidget(add_button)
        form_widget.setLayout(form_layout)

        # Düzenleme ve silme butonları
        button_layout = QHBoxLayout()
        edit_button = QPushButton('Seçili Çalışanı Düzenle')
        edit_button.setIcon(QIcon("docs/icon/employeesedit.png"))
        edit_button.setIconSize(QSize(20, 20))
        edit_button.clicked.connect(self.edit_selected_employee)
        delete_button = QPushButton('Seçili Çalışanı Sil')
        delete_button.setIcon(QIcon("docs/icon/employeesdelete.png"))
        delete_button.setIconSize(QSize(20, 20))
        delete_button.clicked.connect(self.delete_selected_employee)
        
        button_layout.addWidget(edit_button)
        button_layout.addWidget(delete_button)
        
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(6)
        self.employee_table.setHorizontalHeaderLabels(['ID', 'Ad', 'Soyad', 'Pozisyon', 'Maaş', 'İşe Giriş Tarihi'])
        setup_modern_table(self.employee_table)  # Burada fonksiyonu çağırıyoruz

        layout.addWidget(form_widget)
        layout.addLayout(button_layout)
        layout.addWidget(self.employee_table)
        self.setLayout(layout)
        
        self.load_employees()

    def edit_selected_employee(self):
        """Seçili çalışanı düzenle"""
        selected_items = self.employee_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'Uyarı', 'Lütfen düzenlenecek bir çalışan seçin!')
            return
        
        row = selected_items[0].row()
        employee_id = int(self.employee_table.item(row, 0).text())
        self.show_edit_dialog(employee_id)

    def edit_employee(self, item):
        """Tabloda çift tıklanan çalışanı düzenle"""
        row = item.row()
        employee_id = int(self.employee_table.item(row, 0).text())
        self.show_edit_dialog(employee_id)

    






    def show_edit_dialog(self, employee_id):
        """Düzenleme dialog'unu göster"""
        dialog = QDialog(self)
        dialog.setWindowTitle('Çalışan Bilgilerini Düzenle')
        layout = QVBoxLayout()

        # Mevcut çalışan bilgilerini al
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM employees WHERE id = ?', (employee_id,))
        employee = cursor.fetchone()

        # Önceki değişiklikleri gösterecek label
        history_label = QLabel()
        history_label.setStyleSheet("""
            QLabel {
                color: #004d99;
                font-weight: bold;
                padding: 10px;
                background-color: #e6f2ff;
                border-radius: 4px;
            }
        """)

        # Önceki kaydı al ve göster
        cursor.execute('''
            SELECT action_description 
            FROM activity_logs 
            WHERE related_id = ? 
            AND action_type = 'EMPLOYEE_UPDATE'
            ORDER BY timestamp DESC LIMIT 1
        ''', (employee_id,))

        previous_change = cursor.fetchone()
        if previous_change:
            history_label.setText(f"Son Değişiklik:\n{previous_change[0]}")
        history_label.hide()  # Başlangıçta gizle
        layout.addWidget(history_label)

        # Anlık değişiklikleri gösterecek label
        changes_label = QLabel()
        changes_label.setStyleSheet("""
            QLabel {
                color: #CC3300;
                font-weight: bold;
                padding: 5px;
                background-color: #FFE6E6;
                border-radius: 4px;
            }
        """)
        changes_label.hide()
        layout.addWidget(changes_label)

        # Form alanları
        name_input = QLineEdit(employee[1])
        surname_input = QLineEdit(employee[2])
        position_input = QLineEdit(employee[3])
        salary_input = QLineEdit(employee[4])

        # Önceki değerleri kaydet
        original_values = {
            'Ad': employee[1],
            'Soyad': employee[2],
            'Pozisyon': employee[3],
            'Maaş': employee[4]
        }

        # Değişiklikleri takip et
        def show_changes():
            changes = []
            if name_input.text() != original_values['Ad']:
                changes.append(f"Ad: {original_values['Ad']} -> {name_input.text()}")
            if surname_input.text() != original_values['Soyad']:
                changes.append(f"Soyad: {original_values['Soyad']} -> {surname_input.text()}")
            if position_input.text() != original_values['Pozisyon']:
                changes.append(f"Pozisyon: {original_values['Pozisyon']} -> {position_input.text()}")
            if salary_input.text() != original_values['Maaş']:
                changes.append(f"Maaş: {original_values['Maaş']} -> {salary_input.text()}")

            if changes:
                changes_label.setText("Yapılan Değişiklikler:\n" + "\n".join(changes))
                changes_label.show()
                history_label.show()
            else:
                changes_label.hide()
                history_label.hide()

        # Input değişikliklerini dinle
        name_input.textChanged.connect(show_changes)
        surname_input.textChanged.connect(show_changes)
        position_input.textChanged.connect(show_changes)
        salary_input.textChanged.connect(show_changes)

        layout.addWidget(create_label('Ad:', is_form=True))
        layout.addWidget(name_input)
        layout.addWidget(create_label('Soyad:', is_form=True))
        layout.addWidget(surname_input)
        layout.addWidget(create_label('Pozisyon:', is_form=True))
        layout.addWidget(position_input)
        layout.addWidget(create_label('Maaş:', is_form=True))
        layout.addWidget(salary_input)

        # Düzenleme nedeni input
        reason_label = QLabel('Düzenleme Nedeni:')
        reason_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #CC3300;
                margin-top: 10px;
            }
        """)
        layout.addWidget(reason_label)

        reason_input = QLineEdit()
        reason_input.setPlaceholderText('Düzenleme nedenini yazın...')
        reason_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #CC3300;
                border-radius: 4px;
                background-color: #FFFFFF;
            }
            QLineEdit:focus {
                border-color: #E63900;
            }
        """)
        layout.addWidget(reason_input)

        # Kaydet butonu
        button_box = QHBoxLayout()
        save_button = QPushButton('Kaydet')
        save_button.clicked.connect(lambda: self.save_edited_employee(
            dialog, employee_id, name_input.text(), surname_input.text(),
            position_input.text(), salary_input.text(), reason_input.text()
        ))
        cancel_button = QPushButton('İptal')
        cancel_button.clicked.connect(dialog.reject)

        button_box.addWidget(save_button)
        button_box.addWidget(cancel_button)
        layout.addLayout(button_box)

        dialog.setLayout(layout)
        dialog.exec()






    def save_edited_employee(self, dialog, employee_id, name, surname, position, salary, reason=""):
        """Düzenlenen çalışan bilgilerini kaydet"""
        try:
            if not all([name, surname, position, salary]):
                QMessageBox.warning(dialog, 'Hata', 'Tüm alanları doldurun!')
                return
            
            if not reason:
                QMessageBox.warning(dialog, 'Hata', 'Düzenleme nedeni belirtmelisiniz!')
                return
    
            # Önce eski klasör adını almak için mevcut bilgileri al
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT name, surname, position, salary FROM employees WHERE id = ?', (employee_id,))
            old_data = cursor.fetchone()
            old_name, old_surname, old_position, old_salary = old_data
    
            # Maaşı sayısal formata çevir
            salary = float(salary.replace(',', ''))
    
            # Önceki ve yeni durumları hazırla
            previous_state = {
                'Ad': old_name,
                'Soyad': old_surname,
                'Pozisyon': old_position,
                'Maaş': old_salary
            }
            
            new_state = {
                'Ad': name,
                'Soyad': surname,
                'Pozisyon': position,
                'Maaş': f"{salary:,.0f}"
            }
            
            # Değişen alanları belirle
            changes = []
            for key in previous_state:
                if str(previous_state[key]) != str(new_state[key]):
                    changes.append(f"{key}: {previous_state[key]} -> {new_state[key]}")
            
            if not changes:
                QMessageBox.warning(dialog, 'Uyarı', 'Hiçbir değişiklik yapılmadı!')
                return
    
            # Detaylı değişiklik açıklaması oluştur
            changes_description = (
                f"Değişiklikler:\n"
                f"{chr(10).join(changes)}\n\n"
                f"Düzenleme Nedeni: {reason}"
            )
    
            # Veritabanını güncelle
            cursor.execute('''
                UPDATE employees 
                SET name = ?, surname = ?, position = ?, salary = ?
                WHERE id = ?
            ''', (name, surname, position, f"{salary:,.0f}", employee_id))
            
            # Log kaydı ekle
            if self.main_window and self.main_window.current_username:
                self.db.log_activity(
                    self.main_window.current_username,
                    "EMPLOYEE_UPDATE",
                    changes_description,
                    "employees",
                    employee_id
                )
            
            self.db.conn.commit()
    
            # Eski ve yeni klasör yolları
            old_folder_name = f"{old_name}_{old_surname}_{employee_id}"
            new_folder_name = f"{name}_{surname}_{employee_id}"
            old_folder_path = os.path.join(self.file_system.base_dir, old_folder_name)
            new_folder_path = os.path.join(self.file_system.base_dir, new_folder_name)
    
            # Klasör varsa yeniden adlandır
            if os.path.exists(old_folder_path):
                # Excel dosyalarını yeniden adlandır
                for file in os.listdir(old_folder_path):
                    if file.endswith('.xlsx'):
                        old_file_path = os.path.join(old_folder_path, file)
                        new_file = file.replace(f"{old_name}_{old_surname}", f"{name}_{surname}")
                        new_file_path = os.path.join(old_folder_path, new_file)
                        os.rename(old_file_path, new_file_path)
    
                # Klasörü yeniden adlandır
                os.rename(old_folder_path, new_folder_path)
                
            # Tabloyu yenile
            self.load_employees()
            dialog.accept()
            
            self.employee_updated.emit()
            QMessageBox.information(self, 'Başarılı', 'Çalışan bilgileri güncellendi!')
        
        except ValueError:
            QMessageBox.warning(dialog, 'Hata', 'Geçersiz maaş değeri!')
        except Exception as e:
            QMessageBox.warning(dialog, 'Hata', f'Güncelleme hatası: {str(e)}')
    
    def delete_selected_employee(self):
        """Seçili çalışanı sil"""
        selected_items = self.employee_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'Uyarı', 'Lütfen silinecek bir çalışan seçin!')
            return

        row = selected_items[0].row()
        employee_id = int(self.employee_table.item(row, 0).text())
        name = self.employee_table.item(row, 1).text()
        surname = self.employee_table.item(row, 2).text()

        reply = QMessageBox.question(
            self, 'Onay',
            f'{name} {surname} isimli çalışanı silmek istediğinizden emin misiniz?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Önce çalışanın mesai kayıtlarını sil
                cursor = self.db.conn.cursor()
                cursor.execute('DELETE FROM time_records WHERE employee_id = ?', (employee_id,))
                
                # Sonra çalışanı sil
                cursor.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
                
                self.db.conn.commit()
                
                # Tabloyu güncelle
                self.load_employees()
                self.employee_deleted.emit()
                
                QMessageBox.information(self, 'Başarılı', 'Çalışan başarıyla silindi!')
            
            except Exception as e:
                QMessageBox.warning(self, 'Hata', f'Silme işlemi başarısız: {str(e)}')

    def format_salary_input(self, text):
        """Maaş giriş alanını her 3 rakamdan sonra virgül gelecek şekilde formatla"""
        cursor_position = self.salary_input.cursorPosition()
        digit_count = len(text.replace(',', ''))
        
        if digit_count > 3:
            formatted_text = ''
            for i, char in enumerate(reversed(text.replace(',', ''))):
                if i > 0 and i % 3 == 0:
                    formatted_text = ',' + formatted_text
                formatted_text = char + formatted_text
            self.salary_input.setText(formatted_text)
            self.salary_input.setCursorPosition(cursor_position + formatted_text.count(',') - text.count(','))
        elif text != '' and text != ',':
            self.salary_input.setText(text)

    def add_employee(self):
        try:
            name = self.name_input.text().strip()
            surname = self.surname_input.text().strip()
            position = self.position_input.text().strip()
            salary = float(self.salary_input.text().replace(',', ''))

            if not all([name, surname, position]):
                QMessageBox.warning(self, 'Hata', 'Tüm alanları doldurun!')
                return

            # Çalışanı ekle
            employee_id = self.db.add_employee(name, surname, position, salary)
            
            # Log kaydı ekle
            if self.main_window and self.main_window.current_username:
                self.db.log_activity(
                    self.main_window.current_username,
                    "EMPLOYEE_ADD",
                    f"Yeni çalışan eklendi: {name} {surname}",
                    "employees",  # Bu 'Çalışan' olarak görünecek
                    employee_id
                )
            
            try:
                # Çalışan klasörünü oluştur
                folder_path = self.file_system.create_employee_folder(name, surname, employee_id)
                
                QMessageBox.information(
                    self, 'Başarılı', 
                    f'Çalışan başarıyla eklendi!\nKlasör oluşturuldu: {folder_path}'
                )
                
                # Form'u temizle
                for input_field in [self.name_input, self.surname_input, 
                                  self.position_input, self.salary_input]:
                    input_field.clear()

                self.load_employees()
                self.employee_added.emit()
                
            except Exception as e:
                QMessageBox.warning(
                    self, 'Dosya Hatası',
                    f'Çalışan eklendi ancak klasör oluşturulurken hata oluştu: {str(e)}'
                )

        except ValueError:
            QMessageBox.warning(self, 'Hata', 'Geçerli bir maaş değeri girin!')
    
    def load_employees(self):
        employees = self.db.get_all_employees()
        self.employee_table.setRowCount(len(employees))
        
        for row, emp in enumerate(employees):
            for col, value in enumerate(emp):
                item = QTableWidgetItem(str(value))
                if col == 4:  # Maaş sütunu
                    value = value.replace(',', '')  # Virgülleri kaldır
                    item.setData(Qt.ItemDataRole.DisplayRole, float(value))
                    item.setData(Qt.ItemDataRole.EditRole, f"{float(value):,.0f}")
                self.employee_table.setItem(row, col, item)


    
class TimeRecordTab(QWidget):
    time_record_added = pyqtSignal()
    time_record_updated = pyqtSignal()
    time_record_deleted = pyqtSignal()
    def __init__(self, db, main_window=None):  # main_window parametresini ekle
        super().__init__()
        self.db = db
        self.main_window = main_window  # main_window'u kaydet
        self.file_system = EmployeeFileSystem(db=db)

        # Stilleri tanımla
        self.input_style = """
            QLineEdit, QComboBox, QDateEdit {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                min-height: 30px;
                background-color: white;
            }
        """

        # Stilleri tanımla
        self.input_style = """
            QLineEdit, QComboBox, QDateEdit {
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 4px;
                min-height: 30px;
                background-color: white;
            }
        """

        self.button_style = """
            QPushButton {
                background-color: #CC3300;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #E63900;
            }
        """

        self.label_style = """
            QLabel {
                font-weight: bold;
                color: #333;
            }
        """

        self.group_style = """
            QGroupBox {
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 5px;
                background-color: white;
            }
        """

        self.calendar_style = """
            /* Ana takvim widget'ı */
            QCalendarWidget {
                background-color: white;
                font-family: 'Segoe UI', Arial;
                min-width: 200px;
                min-height: 230px;
                selection-background-color: #CC3300;
                selection-color: white;
            }

            /* Takvim başlığı (ay ve yıl) */
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #FFFFFF;
                padding: 5px;
            }

            /* Ay/yıl seçim butonları */
            QCalendarWidget QToolButton {
                color: #333333;
                background-color: transparent;
                font-size: 16px;
                font-weight: bold;
                padding: 8px 5px;
                border-radius: 5px;
            }

            QCalendarWidget QToolButton:hover {
                background-color: #f0f0f0;
            }

            /* Takvim içindeki tüm günler */
            QCalendarWidget QAbstractItemView:enabled {
                font-size: 14px;
                font-weight: bold;
                background-color: white;
                selection-background-color: #CC3300;
                selection-color: white;
                outline: none;
            }

            /* Günlerin başlıkları (Pzt, Sal, vb.) */
            QCalendarWidget QTableView {
                background-color: white;
                font-size: 14px;
                font-weight: bold;
            }

            /* Seçili olmayan günler */
            QCalendarWidget QTableView::item:!selected {
                border-radius: 5px;
                padding: 5px;
            }

            /* Seçili gün */
            QCalendarWidget QTableView::item:selected {
                background-color: #CC3300;
                border-radius: 5px;
                color: white;
                font-weight: bold;
            }

            /* Mouse ile üzerine gelinen gün */
            QCalendarWidget QTableView::item:hover {
                background-color: #f5f5f5;
                border-radius: 5px;
            }

            /* Ay dışı günler */
            QCalendarWidget QAbstractItemView:disabled {
                color: #CCCCCC;
            }
        """

        self.table_style = """
            QTableWidget {
                background-color: white;
                alternate-background-color: #f5f5f5;
                border: 1px solid #ccc;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #CC3300;
                color: white;
                padding: 8px;
                border: none;
            }
        """

        self.checkbox_style = """
            QCheckBox {
                spacing: 20px;
            }
            QCheckBox::indicator {
                width: 40px;
                height: 40px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #ccc;
                border-radius: 4px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #CC3300;
                border-radius: 4px;
                background-color: #CC3300;
            }
        """

        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(10)

        # Üst bölüm için container
        top_container = QWidget()
        top_layout = QHBoxLayout(top_container)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Sol form grubu (%50 genişlik)
        left_group = QGroupBox()
        left_layout = QVBoxLayout(left_group)
        left_layout.setSpacing(10)

        # Sol form elemanları
        self.employee_select = QComboBox()
        self.check_in_input = QLineEdit()
        self.check_in_input.setPlaceholderText('Giriş Saati (HH:MM)')
        self.check_out_input = QLineEdit()
        self.check_out_input.setPlaceholderText('Çıkış Saati (HH:MM)')
        


        # Holiday container widget'ını güncelle
        holiday_container = QWidget()
        holiday_layout = QHBoxLayout()
        holiday_layout.setContentsMargins(0, 10, 0, 10)
        holiday_layout.setSpacing(10)

        # Checkbox'ları oluştur
        self.holiday_checkbox = QCheckBox("Tatil/Bayram Günü")
        self.sick_leave_checkbox = QCheckBox("Raporlu")  # Raporlu günler için checkbox
        self.paid_leave_checkbox = QCheckBox("İzinli")   # İzinli günler için checkbox
        self.weekend_leave_checkbox = QCheckBox("Haftalık izin")  # Hafta sonu izni için checkbox

        # Checkbox'ları birbirini etkisiz hale getirmek için bağlantılar
        self.holiday_checkbox.stateChanged.connect(lambda: self.handle_checkbox_state("holiday"))
        self.sick_leave_checkbox.stateChanged.connect(lambda: self.handle_checkbox_state("sick"))
        self.paid_leave_checkbox.stateChanged.connect(lambda: self.handle_checkbox_state("paid"))

        # Checkbox'lar için ortak stil
        checkbox_style = """
            QCheckBox {
                spacing: 10px;
                font-size: 14px;
                font-weight: bold;
                color: #333333;
            }

            QCheckBox::indicator {
                width: 25px;
                height: 25px;
                border: 2px solid #CC3300;
                border-radius: 6px;
                background-color: white;
            }

            QCheckBox::indicator:unchecked:hover {
                background-color: #FFE6E6;
                border-color: #E63900;
            }

            QCheckBox::indicator:checked {
                background-color: #CC3300;
                border-color: #CC3300;
            }

            QCheckBox::indicator:checked:hover {
                background-color: #E63900;
                border-color: #E63900;
            }

            QCheckBox:hover {
                color: #CC3300;
            }
        """

        # Stilleri uygula
        self.holiday_checkbox.setStyleSheet(checkbox_style)
        self.sick_leave_checkbox.setStyleSheet(checkbox_style)
        self.paid_leave_checkbox.setStyleSheet(checkbox_style)
        self.weekend_leave_checkbox.setStyleSheet(checkbox_style)

        # Checkbox'ları layout'a ekle
        holiday_layout.addWidget(self.holiday_checkbox)
        holiday_layout.addWidget(self.sick_leave_checkbox)
        holiday_layout.addWidget(self.paid_leave_checkbox)
        holiday_layout.addWidget(self.weekend_leave_checkbox)
        holiday_layout.addStretch()
        holiday_container.setLayout(holiday_layout)





        left_layout.addWidget(QLabel('Çalışan:'))
        left_layout.addWidget(self.employee_select)
        left_layout.addWidget(QLabel('Giriş Saati:'))
        left_layout.addWidget(self.check_in_input)
        left_layout.addWidget(QLabel('Çıkış Saati:'))
        left_layout.addWidget(self.check_out_input)
        left_layout.addWidget(holiday_container)
        left_layout.addStretch()

        # Sağ form grubu (%50 genişlik)
        right_group = QGroupBox()
        right_layout = QVBoxLayout(right_group)
        right_layout.setSpacing(10)

        # Sağ form elemanları
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDateTime(QDateTime.currentDateTime())
        self.date_edit.dateChanged.connect(self.check_saturday)

        # Kalıcı takvim widget'ı
        # Takvim widget'ını yapılandır
        self.calendar = self.date_edit.calendarWidget()
        self.calendar.setStyleSheet(self.calendar_style)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)  # Hafta numaralarını kaldır
        self.calendar.setHorizontalHeaderFormat(QCalendarWidget.HorizontalHeaderFormat.ShortDayNames)  # Kısa gün isimleri
        self.calendar.setGridVisible(False)  # Grid çizgilerini kaldır
        self.calendar.setMinimumSize(400, 400)  # Minimum boyut
        self.calendar.setSelectionMode(QCalendarWidget.SelectionMode.SingleSelection)  # Tek gün seçimi
        self.calendar.setNavigationBarVisible(True)  # Gezinme çubuğunu göster

        # YENİ: Tarih değişikliğini dinleme ekle
        self.calendar.selectionChanged.connect(self.on_date_selected)
        self.date_edit.dateChanged.connect(self.on_date_selected)
        # Font ayarları
        palette = self.calendar.palette()
        palette.setColor(palette.ColorRole.WindowText, QColor("#333333"))  # Ana metin rengi
        self.calendar.setPalette(palette)

        self.saturday_label = QLabel()
        self.saturday_label.setStyleSheet("color: blue; font-weight: bold;")

        right_layout.addWidget(QLabel('Tarih:'))
        right_layout.addWidget(self.date_edit)
        right_layout.addWidget(self.calendar)  # Kalıcı takvimi ekle
        right_layout.addWidget(self.saturday_label)
        right_layout.addStretch()

        # Sol ve sağ grupları üst layout'a ekle (genişlikleri %50-%50)
        top_layout.addWidget(left_group, 1)  # stretch factor 1
        top_layout.addWidget(right_group, 1) # stretch factor 1

        # Buton bölümü
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 10, 0, 10)

        record_button = QPushButton('Mesai Kaydı Ekle')
        record_button.clicked.connect(self.add_time_record)
        record_button.setIcon(QIcon("docs/icon/timemanagement.png"))
        record_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        record_button.setMinimumHeight(40)

        edit_button = QPushButton('Seçili Kaydı Düzenle')
        edit_button.setIcon(QIcon("docs/icon/timeedit.png"))
        edit_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        edit_button.clicked.connect(self.edit_selected_record)
        edit_button.setMinimumHeight(40)

        button_layout.addWidget(record_button)
        button_layout.addWidget(edit_button)

        # Tablo bölümü
        self.time_records_table = QTableWidget()
        self.time_records_table.setColumnCount(9)
        self.time_records_table.setHorizontalHeaderLabels([
            'ID', 'Tarih', 'Giriş', 'Çıkış', 'Çalışma Süresi', 
            'Mesai Durumu', 'Mesai Tipi', 'Mesai Ücreti', 'Günlük Ücret'
        ])
        self.time_records_table.itemDoubleClicked.connect(self.edit_time_record)

        # Tablo ayarları
        header = self.time_records_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.time_records_table.setAlternatingRowColors(True)

        # Ana layout'a bölümleri ekle
        main_layout.addWidget(top_container, 35)    # %35 yükseklik
        main_layout.addWidget(button_container, 15) # %15 yükseklik
        main_layout.addWidget(self.time_records_table, 50) # %50 yükseklik

        self.setLayout(main_layout)
        self.update_employee_select()
        self.check_saturday()

    
    
       

    def on_date_selected(self):
        """Seçili tarih değiştiğinde çağrılır"""
        employee_id = self.employee_select.currentData()
        if employee_id:
            self.load_time_records(employee_id)

    def check_saturday(self):
        """Seçilen tarih cumartesi mi kontrol et"""
        selected_date = self.date_edit.date().toPyDate()
        if selected_date.weekday() == 5:  # 5 = Cumartesi
            self.saturday_label.setText("⚠️ Cumartesi günü seçildi (5 saat standart mesai)")
        else:
            self.saturday_label.setText("")

    def add_time_record(self):
        try:
            employee_id = self.employee_select.currentData()
            date = self.date_edit.date().toPyDate().strftime('%Y-%m-%d')

            # Checkbox durumlarını kontrol et
            is_holiday = self.holiday_checkbox.isChecked()  
            is_sick_leave = self.sick_leave_checkbox.isChecked()  
            is_paid_leave = self.paid_leave_checkbox.isChecked()
            is_weekend_leave = self.weekend_leave_checkbox.isChecked()
            is_weekend = self.db.is_weekend(date)

            if not employee_id:
                QMessageBox.warning(self, 'Hata', 'Lütfen çalışan seçin!')
                return

            if self.check_existing_record(employee_id, date):
                QMessageBox.warning(self, 'Hata', 'Bu çalışan için seçilen tarihte zaten kayıt var!')
                return

            # Mesai tipini ve saatlerini belirle
            if is_paid_leave or is_weekend_leave:
                check_in = "00:00"
                check_out = "00:00" 
                working_hours = 0
                overtime_hours = 0
                base_pay = 0
                overtime_pay = 0

                # Haftalık izin için maaş hesaplaması
                if is_weekend_leave:
                    cursor = self.db.conn.cursor()
                    cursor.execute('SELECT salary FROM employees WHERE id = ?', (employee_id,))
                    monthly_salary = cursor.fetchone()[0]
                    monthly_salary = float(monthly_salary.replace(',', ''))
                    daily_rate = monthly_salary / 30  # Aylık maaşı 30 güne böl
                    total_pay = daily_rate
                else:
                    total_pay = 0

            elif is_sick_leave:
                check_in = "09:00"
                check_out = "18:00"
                working_hours = 8
                overtime_hours = 0
                base_pay = 0
                overtime_pay = 0
                total_pay = 0

            else:
                check_in = self.check_in_input.text()
                check_out = self.check_out_input.text()

                if not (check_in and check_out):
                    QMessageBox.warning(self, 'Hata', 'Giriş/çıkış saatlerini girin!')
                    return

                working_hours, overtime_hours = self.db.calculate_working_hours(check_in, check_out, date)
                base_pay, overtime_pay, total_pay = self.db.calculate_pay(
                    employee_id, working_hours, overtime_hours, is_holiday, is_weekend
                )

                if is_holiday:
                    self.db.add_holiday(date, "Manuel Eklenen Tatil")

            # Veritabanına kaydet
            cursor = self.db.conn.cursor()
            cursor.execute('''
            INSERT INTO time_records (
                employee_id, date, check_in, check_out, working_hours, 
                overtime_hours, is_holiday, is_weekend, base_pay, overtime_pay, total_pay,
                is_sick_leave, is_paid_leave, is_weekend_leave
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                employee_id, date, check_in, check_out, working_hours, 
                overtime_hours, is_holiday, is_weekend, base_pay, overtime_pay, total_pay,
                is_sick_leave, is_paid_leave, is_weekend_leave
            ))

            record_id = cursor.lastrowid
            self.db.conn.commit()

            # Log kaydı
            if self.main_window and self.main_window.current_username:
                cursor.execute('SELECT name, surname FROM employees WHERE id = ?', (employee_id,))
                employee = cursor.fetchone()
                if employee:
                    name, surname = employee
                    self.db.log_activity(
                        self.main_window.current_username,
                        "TIME_RECORD_ADD",
                        f"Mesai kaydı eklendi: {name} {surname} - {date}",
                        "time_records",
                        record_id
                    )

            # Excel güncellemesi için veriyi hazırla
            record_data = {
                'date': date,
                'check_in': check_in,
                'check_out': check_out,
                'working_hours': working_hours,
                'regular_hours': working_hours,
                'overtime_hours': overtime_hours,
                'missing_hours': 0,
                'work_type': 'Haftalık İzin' if is_weekend_leave else (
                    'Raporlu' if is_sick_leave else (
                    'İzinli' if is_paid_leave else (
                    'Tatil' if is_holiday else 'Normal'))),
                'regular_pay': base_pay,
                'overtime_pay': overtime_pay,
                'missing_pay': 0,
                'total_pay': total_pay
            }

            # Excel dosyasını güncelle
            if employee:
                if self.file_system.update_employee_timesheet(employee_id, name, surname, record_data):
                    self.check_in_input.clear()
                    self.check_out_input.clear()
                    self.holiday_checkbox.setChecked(False)
                    self.sick_leave_checkbox.setChecked(False)
                    self.paid_leave_checkbox.setChecked(False)
                    self.weekend_leave_checkbox.setChecked(False)
                    self.load_time_records(employee_id)
                    QMessageBox.information(self, 'Başarılı', 'Mesai kaydı başarıyla eklendi!')
                else:
                    QMessageBox.warning(self, 'Excel Güncelleme Hatası', 'Mesai kaydı eklendi ancak Excel dosyası güncellenemedi!')

            self.time_record_added.emit()

        except Exception as e:
            QMessageBox.warning(self, 'Hata', str(e))


    def update_employee(self, employee_id, name, surname, position, salary):
        """Çalışan bilgilerini güncelle"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                UPDATE employees 
                SET name = ?, surname = ?, position = ?, salary = ?
                WHERE id = ?
            ''', (name, surname, position, salary, employee_id))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Güncelleme hatası: {str(e)}")
            return False
        
    def delete_employee(self, employee_id):
        """Çalışanı ve ilişkili tüm kayıtlarını sil"""
        try:
            cursor = self.conn.cursor()
            # Önce çalışanın mesai kayıtlarını sil
            cursor.execute('DELETE FROM time_records WHERE employee_id = ?', (employee_id,))
            # Sonra çalışanı sil
            cursor.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Silme hatası: {str(e)}")
            return False        

    def load_time_records(self, employee_id):
        if not employee_id:
            return

        selected_date = self.date_edit.date().toPyDate()

        # Sadece seçili güne ait kayıtları al
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT * FROM time_records 
            WHERE employee_id = ? AND date = ?
            ORDER BY check_in
        ''', (employee_id, selected_date.strftime('%Y-%m-%d')))

        records = cursor.fetchall()
        self.time_records_table.setRowCount(len(records))

        for row, record in enumerate(records):
            standard_hours = 5 if record[8] else 8  # Cumartesi veya normal gün kontrolü

            if record[12]:  # is_sick_leave
                mesai_tipi = "Raporlu"
                mesai_durumu = "Raporlu"
            elif record[13]:  # is_paid_leave
                mesai_tipi = "İzinli"
                mesai_durumu = "İzinli"
            elif record[14]:  # is_weekend_leave
                mesai_tipi = "Haftalık İzin"
                mesai_durumu = "Haftalık İzin"
            elif record[7]:  # is_holiday
                mesai_tipi = "Tatil"
                mesai_durumu = "Tatil"
            else:
                mesai_tipi = "Normal"
                if record[5] < standard_hours:
                    eksik_saat = standard_hours - record[5]
                    mesai_durumu = f"Eksik ({format_time_as_hours_minutes(eksik_saat)})"
                else:
                    if record[6] > 0:  # Fazla mesai varsa
                        mesai_durumu = f"Fazla mesai: {format_time_as_hours_minutes(record[6])}"
                    else:
                        mesai_durumu = "Tam mesai"

            # Tablo hücrelerini doldur
            self.time_records_table.setItem(row, 0, QTableWidgetItem(str(record[0])))  # ID
            self.time_records_table.setItem(row, 1, QTableWidgetItem(record[2]))       # Tarih
            self.time_records_table.setItem(row, 2, QTableWidgetItem(record[3]))       # Giriş
            self.time_records_table.setItem(row, 3, QTableWidgetItem(record[4]))       # Çıkış
            self.time_records_table.setItem(row, 4, QTableWidgetItem(format_time_as_hours_minutes(record[5])))  # Çalışma Süresi
            self.time_records_table.setItem(row, 5, QTableWidgetItem(mesai_durumu))    # Mesai Durumu
            self.time_records_table.setItem(row, 6, QTableWidgetItem(mesai_tipi))      # Mesai Tipi

            # Muhasebeci için ücret bilgilerini gizle
            if self.main_window and self.main_window.user_permissions.get('department') == 'Muhasebe':
                self.time_records_table.setItem(row, 7, QTableWidgetItem("-"))    # Mesai Ücreti
                self.time_records_table.setItem(row, 8, QTableWidgetItem("-"))    # Günlük Ücret
            else:
                self.time_records_table.setItem(row, 7, QTableWidgetItem(f"{record[10]:.0f} TL"))    # Mesai Ücreti
                self.time_records_table.setItem(row, 8, QTableWidgetItem(f"{record[11]:.0f} TL"))    # Günlük Ücret

            # Eksik mesai satırını kırmızı yap
            if record[5] < standard_hours:
                for col in range(self.time_records_table.columnCount()):
                    item = self.time_records_table.item(row, col)
                    if item:
                        item.setForeground(Qt.GlobalColor.red)


    def edit_selected_record(self):
        """Seçili kaydı düzenle"""
        selected_items = self.time_records_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, 'Uyarı', 'Lütfen düzenlenecek bir kayıt seçin!')
            return
            
        row = selected_items[0].row()
        record_id = int(self.time_records_table.item(row, 0).text())
        self.show_edit_dialog(record_id, row)

    def edit_time_record(self, item):
        """Tabloda çift tıklanan kaydı düzenle"""
        record_id = int(self.time_records_table.item(item.row(), 0).text())
        self.show_edit_dialog(record_id, item.row())

    def show_edit_dialog(self, record_id, row):
        """Düzenleme dialog'unu göster"""
        dialog = QDialog(self)
        dialog.setWindowTitle('Mesai Kaydını Düzenle')
        layout = QVBoxLayout()

        # Mevcut değerleri al
        current_check_in = self.time_records_table.item(row, 2).text()
        current_check_out = self.time_records_table.item(row, 3).text()

        # Form container
        form_container = QWidget()
        form_layout = QFormLayout()

        # Giriş alanları
        check_in_input = QLineEdit(current_check_in)
        check_out_input = QLineEdit(current_check_out)

        check_in_input.setPlaceholderText('HH:MM')
        check_out_input.setPlaceholderText('HH:MM')

        # Düzenleme nedeni için input alanı
        reason_input = QLineEdit()
        reason_input.setPlaceholderText('Düzenleme nedenini yazın...')
        reason_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #CC3300;
                border-radius: 4px;
                background-color: #FFFFFF;
                min-height: 30px;
            }
            QLineEdit:focus {
                border-color: #E63900;
            }
        """)

        form_layout.addRow(create_label('Giriş Saati (HH:MM):', is_form=True), check_in_input)
        form_layout.addRow(create_label('Çıkış Saati (HH:MM):', is_form=True), check_out_input)
        form_layout.addRow(create_label('Düzenleme Nedeni:', is_form=True), reason_input)

        # Butonlar için horizontal layout
        button_layout = QHBoxLayout()

        # Kaydet butonu
        save_button = QPushButton('Kaydet')
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_button.clicked.connect(lambda: self.save_edited_record(
            dialog, record_id, check_in_input.text(), check_out_input.text(), reason_input.text()
        ))

        # İptal butonu
        cancel_button = QPushButton('İptal')
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        cancel_button.clicked.connect(dialog.reject)

        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)

        # Form container'ı layout'a ekle
        form_layout.addRow(button_layout)
        form_container.setLayout(form_layout)
        layout.addWidget(form_container)

        dialog.setLayout(layout)
        dialog.exec()

    def save_edited_record(self, dialog, record_id, check_in, check_out, reason=""):
        """Düzenlenen kaydı kaydet"""
        try:
            if not all([check_in, check_out]):
                QMessageBox.warning(dialog, 'Hata', 'Giriş ve çıkış saatlerini girin!')
                return
                
            if not reason:
                QMessageBox.warning(dialog, 'Hata', 'Düzenleme nedeni gereklidir!')
                return
    
            # Önceki kaydı al
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT t.*, e.name, e.surname
                FROM time_records t
                JOIN employees e ON t.employee_id = e.id
                WHERE t.id = ?
            ''', (record_id,))
            old_record = cursor.fetchone()
    
            if not old_record:
                QMessageBox.warning(self, 'Hata', 'Kayıt bulunamadı!')
                return
    
            # Değişiklikleri belirle
            changes = []
            if old_record[3] != check_in:
                changes.append(f"Giriş Saati: {old_record[3]} -> {check_in}")
            if old_record[4] != check_out:
                changes.append(f"Çıkış Saati: {old_record[4]} -> {check_out}")
    
            if not changes:
                QMessageBox.warning(self, 'Uyarı', 'Hiçbir değişiklik yapılmadı!')
                return
    
            # Yeni çalışma saatlerini hesapla
            working_hours, overtime_hours = self.db.calculate_working_hours(check_in, check_out, old_record[2])
            is_weekend = self.db.is_weekend(old_record[2])
    
            # Yeni ücretleri hesapla
            base_pay, overtime_pay, total_pay = self.db.calculate_pay(
                old_record[1], working_hours, overtime_hours, old_record[7], is_weekend
            )
    
            # Önceki ve yeni durumların detaylı karşılaştırması
            status_comparison = (
                f"Önceki Durum:\n"
                f"Çalışma Süresi: {old_record[5]:.1f} saat\n"
                f"Fazla Mesai: {old_record[6]:.1f} saat\n"
                f"Ücret: {old_record[11]:.2f} TL\n\n"
                f"Yeni Durum:\n"
                f"Çalışma Süresi: {working_hours:.1f} saat\n"
                f"Fazla Mesai: {overtime_hours:.1f} saat\n"
                f"Ücret: {total_pay:.2f} TL"
            )
    
            # Log kaydı için detaylı açıklama oluştur
            changes_description = (
                f"Çalışan: {old_record[-2]} {old_record[-1]}\n"
                f"Tarih: {old_record[2]}\n\n"
                f"Yapılan Değişiklikler:\n"
                f"{chr(10).join(changes)}\n\n"
                f"{status_comparison}\n\n"
                f"Düzenleme Nedeni: {reason}"
            )
    
            # Veritabanını güncelle
            cursor.execute('''
                UPDATE time_records 
                SET check_in = ?, check_out = ?, working_hours = ?, overtime_hours = ?,
                    base_pay = ?, overtime_pay = ?, total_pay = ?
                WHERE id = ?
            ''', (check_in, check_out, working_hours, overtime_hours,
                  base_pay, overtime_pay, total_pay, record_id))
            
            # Excel dosyasını güncelle
            try:
                record_data = {
                    'date': old_record[2],
                    'check_in': check_in,
                    'check_out': check_out,
                    'working_hours': working_hours,
                    'regular_hours': working_hours - overtime_hours,
                    'overtime_hours': overtime_hours,
                    'missing_hours': max(0, 8 - working_hours) if not is_weekend else max(0, 5 - working_hours),
                    'work_type': 'Haftalık İzin' if old_record[14] else (
                        'Raporlu' if old_record[12] else (
                        'İzinli' if old_record[13] else (
                        'Tatil' if old_record[7] else 'Normal'))),
                    'regular_pay': base_pay,
                    'overtime_pay': overtime_pay,
                    'missing_pay': 0,
                    'total_pay': total_pay
                }
    
                file_system = EmployeeFileSystem(db=self.db)
                if not file_system.update_employee_timesheet(old_record[1], old_record[-2], old_record[-1], record_data):
                    QMessageBox.warning(self, 'Excel Güncelleme Hatası', 
                                      'Mesai kaydı güncellendi ancak Excel dosyası güncellenemedi!')
    
            except Exception as excel_error:
                print(f"Excel güncelleme hatası: {str(excel_error)}")
                QMessageBox.warning(self, 'Excel Hatası', 
                                  'Mesai kaydı güncellendi ancak Excel dosyası güncellenirken hata oluştu!')
            
            # Log kaydı ekle
            if self.main_window and self.main_window.current_username:
                self.db.log_activity(
                    self.main_window.current_username,
                    "TIME_RECORD_UPDATE",
                    changes_description,
                    "time_records",
                    record_id
                )
    
            self.db.conn.commit()
            dialog.accept()
            self.load_time_records(old_record[1])  # employee_id ile kayıtları yenile
            QMessageBox.information(self, 'Başarılı', 'Kayıt güncellendi!')
            
            self.time_record_updated.emit()
    
        except ValueError:
            QMessageBox.warning(self, 'Hata', 'Geçersiz saat formatı! Lütfen HH:MM formatında girin.')
        except Exception as e:
            QMessageBox.warning(self, 'Hata', f'Güncelleme hatası: {str(e)}')

    def update_employee_select(self):
        """Çalışan seçim listesini güncelle"""
        current_employee = self.employee_select.currentData() if self.employee_select.count() > 0 else None
        
        self.employee_select.clear()
        employees = self.db.get_all_employees()
        for emp in employees:
            self.employee_select.addItem(f"{emp[1]} {emp[2]} (ID: {emp[0]})", emp[0])
            
        # Önceki seçili çalışanı tekrar seç
        if current_employee:
            index = self.employee_select.findData(current_employee)
            if index >= 0:
                self.employee_select.setCurrentIndex(index)
                self.load_time_records(current_employee)

    def check_existing_record(self, employee_id, date):
        """Belirli bir tarihte çalışanın kaydı var mı kontrol et"""
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT id FROM time_records 
            WHERE employee_id = ? AND date = ?
        ''', (employee_id, date))
        return cursor.fetchone() is not None
    

    def handle_checkbox_state(self, checkbox_type):
        """
        Checkbox'ların durumlarını yönetir ve birbirlerini etkisizleştirir

        Args:
            checkbox_type (str): Tıklanan checkbox'ın tipi ('holiday', 'sick', 'paid', veya 'weekend')
        """
        if checkbox_type == "holiday" and self.holiday_checkbox.isChecked():
            self.sick_leave_checkbox.setChecked(False)
            self.paid_leave_checkbox.setChecked(False)
            self.weekend_leave_checkbox.setChecked(False)
        elif checkbox_type == "sick" and self.sick_leave_checkbox.isChecked():
            self.holiday_checkbox.setChecked(False)
            self.paid_leave_checkbox.setChecked(False)
            self.weekend_leave_checkbox.setChecked(False)
        elif checkbox_type == "paid" and self.paid_leave_checkbox.isChecked():
            self.holiday_checkbox.setChecked(False)
            self.sick_leave_checkbox.setChecked(False)
            self.weekend_leave_checkbox.setChecked(False)
        elif checkbox_type == "weekend" and self.weekend_leave_checkbox.isChecked():
            self.holiday_checkbox.setChecked(False)
            self.sick_leave_checkbox.setChecked(False)
            self.paid_leave_checkbox.setChecked(False)

class ReportingTab(QWidget):
    report_generated = pyqtSignal()
    report_updated = pyqtSignal()
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.init_ui()

    def generate_pdf_report(self):
        """PDF rapor oluşturma metodu"""
        try:
            # Seçili çalışan ve tarih bilgilerini al
            employee_id = self.employee_select.currentData()
            month = self.month_select.currentData()
            year = self.year_select.currentData()
            
            # Çalışan seçili değilse uyarı ver ve çık
            if not employee_id:
                QMessageBox.warning(self, 'Uyarı', 'Lütfen bir çalışan seçin!')
                return
            
            try:    
                # Veritabanından gerekli bilgileri çek
                employee = self.db.get_employee(employee_id)  # Çalışan bilgileri
                time_records = self.db.get_monthly_records(employee_id, month, year)  # Mesai kayıtları
                summary = self.db.get_employee_monthly_summary(employee_id, month, year)  # Aylık özet
                
                # Veritabanı sorguları başarısız olursa
                if not all([employee, time_records, summary]):
                    QMessageBox.warning(self, 'Veri Hatası', 'Çalışan bilgileri veya kayıtları alınamadı!')
                    return
                    
                # PDF oluşturma işlemi
                pdf_generator = PDFGenerator()
                pdf_path = pdf_generator.generate_monthly_report(
                    employee_data=employee,
                    time_records=time_records,
                    summary_data=summary,
                    month=month,
                    year=year
                )
                
                # PDF başarıyla oluşturulduysa
                if pdf_path and os.path.exists(pdf_path):
                    # Başarı mesajı göster
                    QMessageBox.information(
                        self, 
                        'Başarılı', 
                        f'PDF rapor başarıyla oluşturuldu!\nDosya konumu: {pdf_path}'
                    )
                    
                    # PDF'i varsayılan uygulamada aç
                    if os.name == 'nt':  # Windows işletim sistemi
                        os.startfile(pdf_path)
                    else:  # Linux/Mac işletim sistemi
                        os.system(f'xdg-open {pdf_path}')
                        
                    # Başarılı rapor oluşturma sinyali gönder
                    self.report_generated.emit()
                    
                else:
                    QMessageBox.warning(self, 'Hata', 'PDF dosyası oluşturulamadı!')
                    
            except Exception as db_error:
                QMessageBox.critical(self, 'Veritabanı Hatası', 
                                   f'Veri alma işlemi başarısız: {str(db_error)}')
                
        except Exception as e:
            QMessageBox.critical(self, 'Genel Hata', 
                               f'PDF oluşturma işlemi başarısız: {str(e)}')

    def init_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Üst kontrol paneli - daha kompakt
        control_container = QWidget()
        control_container.setMaximumHeight(60)  # Yüksekliği sınırla
        control_layout = QHBoxLayout()
        control_layout.setSpacing(10)
        control_layout.setContentsMargins(10, 5, 10, 5)  # Kenar boşluklarını azalt

        # Çalışan seçimi
        employee_widget = QWidget()
        employee_layout = QHBoxLayout()
        employee_layout.setContentsMargins(0, 0, 0, 0)
        employee_label = QLabel("Çalışan:")
        employee_label.setStyleSheet("font-weight: bold;")
        self.employee_select = QComboBox()
        self.employee_select.setMinimumWidth(200)
        employee_layout.addWidget(employee_label)
        employee_layout.addWidget(self.employee_select)
        employee_widget.setLayout(employee_layout)

        # Ay seçimi
        month_widget = QWidget()
        month_layout = QHBoxLayout()
        month_layout.setContentsMargins(0, 0, 0, 0)
        month_label = QLabel("Ay:")
        month_label.setStyleSheet("font-weight: bold;")
        self.month_select = QComboBox()
        self.month_select.setMinimumWidth(100)
        for i in range(1, 13):
            self.month_select.addItem(f"{i}. Ay", i)
        current_month = datetime.now().month
        self.month_select.setCurrentIndex(current_month - 1)
        month_layout.addWidget(month_label)
        month_layout.addWidget(self.month_select)
        month_widget.setLayout(month_layout)

        # Yıl seçimi
        year_widget = QWidget()
        year_layout = QHBoxLayout()
        year_layout.setContentsMargins(0, 0, 0, 0)
        year_label = QLabel("Yıl:")
        year_label.setStyleSheet("font-weight: bold;")
        self.year_select = QComboBox()
        self.year_select.setMinimumWidth(100)
        current_year = datetime.now().year
        for year in range(2024, current_year + 2):
            self.year_select.addItem(str(year), year)
        year_layout.addWidget(year_label)
        year_layout.addWidget(self.year_select)
        year_widget.setLayout(year_layout)

        # Butonlar
        button_widget = QWidget()
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(5)

        refresh_button = QPushButton("Güncelle")
        refresh_button.setIcon(QIcon("docs/icon/update.png"))
        refresh_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        refresh_button.clicked.connect(self.load_report)
        self.pdf_button = QPushButton("PDF")
        self.pdf_button.setIcon(QIcon("docs/icon/pdf.png"))
        self.pdf_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        self.pdf_button.clicked.connect(self.generate_pdf_report)
        self.excel_button = QPushButton("Excel")
        self.excel_button.setIcon(QIcon("docs/icon/excel.png"))
        self.excel_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        self.excel_button.clicked.connect(self.open_excel_file)

        for btn in [refresh_button, self.pdf_button, self.excel_button]:
            btn.setMinimumWidth(80)
            btn.setMaximumWidth(80)
            btn.setMinimumHeight(30)
            btn.setMaximumHeight(30)

        refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #4A4A4A;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #666666; }
        """)

        self.pdf_button.setStyleSheet("""
            QPushButton {
                background-color: #CC3300;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #E63900; }
        """)

        self.excel_button.setStyleSheet("""
            QPushButton {
                background-color: #217346;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1e8449; }
        """)

        button_layout.addWidget(refresh_button)
        button_layout.addWidget(self.pdf_button)
        button_layout.addWidget(self.excel_button)
        button_widget.setLayout(button_layout)

        # Kontrol elemanlarını layout'a ekle
        control_layout.addWidget(employee_widget)
        control_layout.addWidget(month_widget)
        control_layout.addWidget(year_widget)
        control_layout.addWidget(button_widget)
        control_layout.addStretch()

        control_container.setLayout(control_layout)
        main_layout.addWidget(control_container)

        # Tablolar için container
        tables_container = QWidget()
        tables_layout = QVBoxLayout()
        tables_layout.setSpacing(10)
        tables_layout.setContentsMargins(0, 0, 0, 0)

        # Çalışan bilgileri tablosu
        self.info_table = QTableWidget()
        self.info_table.setColumnCount(2)
        self.info_table.setHorizontalHeaderLabels(['Bilgi', 'Değer'])
        self.info_table.setMaximumHeight(200)
        setup_modern_table(self.info_table)

        # Aylık özet tablosu
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(2)
        self.summary_table.setHorizontalHeaderLabels(['Özet', 'Değer'])
        self.summary_table.setMaximumHeight(200)
        setup_modern_table(self.summary_table)

        # Mesai detayları tablosu
        self.details_table = QTableWidget()
        self.details_table.setColumnCount(7)
        self.details_table.setHorizontalHeaderLabels([
            'Tarih', 'Giriş', 'Çıkış', 'Çalışma Süresi',
            'Mesai Durumu', 'Mesai Tipi', 'Günlük Ücret'
        ])
        setup_modern_table(self.details_table)

        # Tabloları ana container'a ekle
        tables_layout.addWidget(self.info_table)
        tables_layout.addWidget(self.summary_table)
        tables_layout.addWidget(self.details_table, 1)  # Mesai detayları için daha fazla alan
        tables_container.setLayout(tables_layout)

        main_layout.addWidget(tables_container)
        self.setLayout(main_layout)

        # Başlangıç yüklemesi
        self.update_employee_select()
        self.load_report()


    
    def open_excel_file(self):
        """Excel dosyasını aç"""
        try:
            employee_id = self.employee_select.currentData()
            month = self.month_select.currentData()
            year = self.year_select.currentData()
            
            if not employee_id:
                QMessageBox.warning(self, 'Uyarı', 'Lütfen bir çalışan seçin!')
                return
                
            # Çalışan bilgilerini al
            employee = self.db.get_employee(employee_id)
            name, surname = employee[1], employee[2]
            
            # Excel dosya yolunu oluştur
            month_date = datetime(year, month, 1)
            month_name = month_date.strftime('%B')
            excel_name = f"{year}_{month:02d}_{month_name}.xlsx"
            folder_name = f"{name}_{surname}_{employee_id}"
            excel_path = os.path.join('docs/employee_files', folder_name, excel_name)  # Yol güncellendi
            
            # Şifreli dosya varsa deşifrele
            if os.path.exists(excel_path + '.encrypted'):
                try:
                    self.db.security.decrypt_file(excel_path + '.encrypted')
                except Exception as e:
                    QMessageBox.warning(self, 'Hata', f'Dosya şifresi çözülemedi: {str(e)}')
                    return
            
            # Excel dosyası varsa aç
            if os.path.exists(excel_path):
                os.startfile(excel_path) if os.name == 'nt' else os.system(f'xdg-open {excel_path}')
            else:
                QMessageBox.warning(self, 'Hata', 'Excel dosyası bulunamadı!')
                
        except Exception as e:
            QMessageBox.critical(self, 'Hata', f'Excel dosyası açılırken bir hata oluştu: {str(e)}')

    def update_employee_select(self):
        """Çalışan seçim listesini güncelle"""
        current_employee = self.employee_select.currentData() if self.employee_select.count() > 0 else None
        
        self.employee_select.clear()
        employees = self.db.get_all_employees()
        for emp in employees:
            self.employee_select.addItem(f"{emp[1]} {emp[2]} (ID: {emp[0]})", emp[0])
            
        # Önceki seçili çalışanı tekrar seç
        if current_employee:
            index = self.employee_select.findData(current_employee)
            if index >= 0:
                self.employee_select.setCurrentIndex(index)
                self.load_report()

    def load_report(self):
        employee_id = self.employee_select.currentData()
        month = self.month_select.currentData()
        year = self.year_select.currentData()
    
        if not employee_id:
            return
    
        # Çalışan bilgilerini yükle
        employee = self.db.get_employee(employee_id)
        info_data = [
            ('ID', employee[0]),
            ('Ad', employee[1]),
            ('Soyad', employee[2]),
            ('Pozisyon', employee[3]),
            ('Maaş', f"{float(employee[4].replace(',', '')):,.0f} TL"),
            ('İşe Giriş Tarihi', employee[5])
        ]
        self.info_table.setRowCount(len(info_data))
        for row, (key, value) in enumerate(info_data):
            self.info_table.setItem(row, 0, QTableWidgetItem(str(key)))
            self.info_table.setItem(row, 1, QTableWidgetItem(str(value)))
    
       
       





        # Aylık özeti yükle - Bu kısmı güncelliyoruz
        summary = self.db.get_employee_monthly_summary(employee_id, month, year)
        
        if summary and summary[0]:  # Eğer kayıt varsa
            # Tatil günü çalışma sayısını hesapla
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT COUNT(*) 
                FROM time_records 
                WHERE employee_id = ? 
                AND strftime('%m', date) = ? 
                AND strftime('%Y', date) = ?
                AND is_holiday = 1
                AND working_hours > 0
            ''', (employee_id, f"{month:02d}", str(year)))
            tatil_gun_sayisi = cursor.fetchone()[0]
            
            # Özet verileri güncellendi
            summary_data = [
                ('Toplam Çalışılan Gün', summary[0]),
                ('Toplam Çalışma Saati', format_time_as_hours_minutes(summary[1])),
                ('Toplam Fazla Mesai', format_time_as_hours_minutes(summary[2])),
                ('Bayram/Tatil Günü Çalışma', f"{tatil_gun_sayisi} gün"),
                ('Normal Mesai Ücreti', f"{summary[3]:,.2f} TL"),
                ('Fazla Mesai Ücreti', f"{summary[4]:,.2f} TL"),
                ('Toplam Ücret', f"{summary[5]:,.2f} TL")
            ]
        else:  # Kayıt yoksa
            summary_data = [
                ('Toplam Çalışılan Gün', 0),
                ('Toplam Çalışma Saati', '0 saat'),
                ('Toplam Fazla Mesai', '0 saat'),
                ('Bayram/Tatil Günü Çalışma', '0 gün'),
                ('Normal Mesai Ücreti', '0.00 TL'),
                ('Fazla Mesai Ücreti', '0.00 TL'),
                ('Toplam Ücret', '0.00 TL')
            ]

        # Tablo güncelleme kısmı...
        self.summary_table.setRowCount(len(summary_data))
        for row, (key, value) in enumerate(summary_data):
            key_item = QTableWidgetItem(str(key))
            value_item = QTableWidgetItem(str(value))
            
            # Son satırı kalın yap
            if row == len(summary_data) - 1:
                font = key_item.font()
                font.setBold(True)
                key_item.setFont(font)
                value_item.setFont(font)
                key_item.setBackground(Qt.GlobalColor.lightGray)
                value_item.setBackground(Qt.GlobalColor.lightGray)
            
            # Tatil günü çalışma satırını vurgula
            if row == 3:  # Bayram/Tatil satırı
                key_item.setForeground(QColor("#CC3300"))  # Kırmızı renk
                value_item.setForeground(QColor("#CC3300"))
            
            self.summary_table.setItem(row, 0, key_item)
            self.summary_table.setItem(row, 1, value_item)














    
        # Detaylı mesai kayıtlarını yükle
        records = self.db.get_monthly_records(employee_id, month, year)
        self.details_table.setRowCount(len(records))
        for row, record in enumerate(records):
            # Mesai tipini belirle
            if record[14]:  # is_weekend_leave
                mesai_tipi = "Haftalık İzin"
                mesai_durumu = "Haftalık İzin"
            elif record[7]:  # is_holiday
                mesai_tipi = "Tatil"
                mesai_durumu = "Tatil"
            elif record[12]:  # is_sick_leave
                mesai_tipi = "Raporlu"
                mesai_durumu = "Raporlu"
            elif record[13]:  # is_paid_leave
                mesai_tipi = "İzinli"
                mesai_durumu = "İzinli"
            elif record[8]:  # is_weekend
                mesai_tipi = "Cumartesi"
                if record[5] < 5:  # 5 saatten az çalışma
                    mesai_durumu = f"Eksik ({5 - record[5]:.0f} saat)"
                else:
                    mesai_durumu = f"Fazla Mesai: {record[6]:.0f} saat" if record[6] > 0 else "Tam"
            else:
                mesai_tipi = "Normal"
                if record[5] < 8:  # 8 saatten az çalışma
                    mesai_durumu = f"Eksik ({8 - record[5]:.0f} saat)"
                else:
                    mesai_durumu = f"Fazla Mesai: {record[6]:.0f} saat" if record[6] > 0 else "Tam"
    
            # Tabloyu doldur
            self.details_table.setItem(row, 0, QTableWidgetItem(record[2]))  # Tarih
            self.details_table.setItem(row, 1, QTableWidgetItem(record[3]))  # Giriş
            self.details_table.setItem(row, 2, QTableWidgetItem(record[4]))  # Çıkış
            self.details_table.setItem(row, 3, QTableWidgetItem(f"{record[5]:.0f} Saat"))  # Çalışma Süresi
            self.details_table.setItem(row, 4, QTableWidgetItem(mesai_durumu))  # Mesai Durumu
            self.details_table.setItem(row, 5, QTableWidgetItem(mesai_tipi))  # Mesai Tipi
            self.details_table.setItem(row, 6, QTableWidgetItem(f"{record[11]:.0f} TL"))  # Günlük Ücret
    
            # Eksik mesai satırını kırmızı yap (haftalık izin hariç)
            if not record[14] and record[5] < (5 if record[8] else 8):
                for col in range(self.details_table.columnCount()):
                    item = self.details_table.item(row, col)
                    item.setForeground(Qt.GlobalColor.red)
        
    
    def update_year_options(self):
        """Yıl seçeneklerini güncelle"""
        current_year = datetime.now().year
        
        # Mevcut seçili yılı koru
        selected_year = self.year_select.currentData() if self.year_select.count() > 0 else current_year
        
        # Yıl listesini temizle
        self.year_select.clear()
        
        # 2024'ten başlayarak şu anki yıla kadar olan yılları ekle
        start_year = 2024
        for year in range(start_year, current_year + 2):  # Bir sonraki yılı da ekle
            self.year_select.addItem(str(year), year)
            
        # Önceki seçili yılı veya varsayılan olarak güncel yılı seç
        index = self.year_select.findData(selected_year)
        if index >= 0:
            self.year_select.setCurrentIndex(index)

    def check_year_update(self):
        """Yıl seçimi değiştiğinde yeni yıl eklenmesi gerekip gerekmediğini kontrol et"""
        current_year = datetime.now().year
        max_year = max(self.year_select.itemData(i) for i in range(self.year_select.count()))
        
        if current_year + 1 > max_year:
            self.update_year_options()


class CompanyReportTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Tarih seçimi
        date_widget = QWidget()
        date_layout = QHBoxLayout()

        # Her iki sınıfta da init_ui metodunda ayın ve yılın ayarlanması:

        # Tarih seçimi
        self.month_select = QComboBox()
        for i in range(1, 13):
            self.month_select.addItem(f"{i}. Ay", i)
        # Güncel ayı seç
        current_month = datetime.now().month
        self.month_select.setCurrentIndex(current_month - 1)
        
        # Yıl seçimi
        self.year_select = QComboBox()
        current_year = datetime.now().year
        for year in range(2024, current_year + 2):
            self.year_select.addItem(str(year), year)
        # Güncel yılı seç
        current_year_index = self.year_select.findData(current_year)
        if current_year_index >= 0:
            self.year_select.setCurrentIndex(current_year_index)

        refresh_button = QPushButton("Raporu Güncelle")
        refresh_button.setIcon(QIcon("docs/icon/update.png"))
        refresh_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        refresh_button.clicked.connect(self.load_report)

        date_layout.addWidget(create_label("Ay:", is_form=True))
        date_layout.addWidget(self.month_select)
        date_layout.addWidget(create_label("Yıl:", is_form=True))
        date_layout.addWidget(self.year_select)
        date_layout.addWidget(refresh_button)
        date_layout.addStretch()
        date_widget.setLayout(date_layout)
        layout.addWidget(date_widget)

        # Çalışan detayları tablosu
        self.employee_details_table = QTableWidget()
        self.employee_details_table.setColumnCount(8)
        self.employee_details_table.setHorizontalHeaderLabels([
            'ID', 'Ad Soyad', 'Pozisyon', 'Maaş',
            'Toplam Mesai Saati', 'Toplam Mesai Ücreti',
            'Toplam Ödenecek', 'Durum'
        ])
        layout.addWidget(create_label("Çalışan Detayları:", is_form=True))
        layout.addWidget(self.employee_details_table)
        setup_modern_table(self.employee_details_table)

        # İşletme toplam tablosu
        self.company_total_table = QTableWidget()
        self.company_total_table.setColumnCount(2)
        self.company_total_table.setHorizontalHeaderLabels(['Bilgi', 'Tutar'])
        layout.addWidget(create_label("İşletme Toplam Ödemeleri:", is_form=True))
        layout.addWidget(self.company_total_table)
        setup_modern_table(self.employee_details_table)

        self.setLayout(layout)
        self.load_report()

    def get_employee_monthly_data(self, employee_id, month, year):
        """Çalışanın aylık verilerini hesapla"""
        cursor = self.db.conn.cursor()
        
        # Mesai bilgilerini al
        cursor.execute('''
            SELECT 
                SUM(working_hours) as total_hours,
                SUM(overtime_hours) as total_overtime,
                SUM(overtime_pay) as total_overtime_pay
            FROM time_records 
            WHERE employee_id = ? 
            AND strftime('%m', date) = ? 
            AND strftime('%Y', date) = ?
        ''', (employee_id, f"{month:02d}", str(year)))
        
        result = cursor.fetchone()
        return {
            'total_hours': result[0] or 0,
            'total_overtime': result[1] or 0,
            'total_overtime_pay': result[2] or 0
        }

    def load_report(self):
        month = self.month_select.currentData()
        year = self.year_select.currentData()

        # Tüm çalışanları al
        employees = self.db.get_all_employees()
        self.employee_details_table.setRowCount(len(employees))

        total_salary = 0
        total_overtime_pay = 0
        total_payment = 0

        for row, emp in enumerate(employees):
            emp_id, name, surname, position, salary, hire_date = emp
            salary = float(salary.replace(',', ''))  # Maaşı float'a dönüştür
            monthly_data = self.get_employee_monthly_data(emp_id, month, year)

            total = salary + monthly_data['total_overtime_pay']

            # Durum (Mesai yapıp yapmadığı)
            status = "Mesai Var" if monthly_data['total_overtime'] > 0 else "Mesai Yok"
            if monthly_data['total_hours'] == 0:
                status = "Çalışma Yok"

            # Tabloya verileri ekle
            self.employee_details_table.setItem(row, 0, QTableWidgetItem(str(emp_id)))
            self.employee_details_table.setItem(row, 1, QTableWidgetItem(f"{name} {surname}"))
            self.employee_details_table.setItem(row, 2, QTableWidgetItem(position))
            self.employee_details_table.setItem(row, 3, QTableWidgetItem(f"{salary:,.0f} TL"))
            self.employee_details_table.setItem(row, 4, QTableWidgetItem(f"{monthly_data['total_hours']:.0f} Saat"))
            self.employee_details_table.setItem(row, 5, QTableWidgetItem(f"{monthly_data['total_overtime_pay']:,.0f} TL"))
            self.employee_details_table.setItem(row, 6, QTableWidgetItem(f"{total:,.0f} TL"))
            self.employee_details_table.setItem(row, 7, QTableWidgetItem(status))

            # Toplamları güncelle
            total_salary += salary
            total_overtime_pay += monthly_data['total_overtime_pay']
            total_payment += total

        # İşletme toplamlarını göster
        company_data = [
            ('Toplam Maaş Ödemeleri', f"{total_salary:,.0f} TL"),
            ('Toplam Mesai Ödemeleri', f"{total_overtime_pay:,.0f} TL"),
            ('TOPLAM ÖDEME', f"{total_payment:,.0f} TL")
        ]

        self.company_total_table.setRowCount(len(company_data))
        for row, (key, value) in enumerate(company_data):
            key_item = QTableWidgetItem(key)
            value_item = QTableWidgetItem(value)
            
            # Son satırı kalın ve renkli yap
            if row == len(company_data) - 1:
                font = key_item.font()
                font.setBold(True)
                key_item.setFont(font)
                value_item.setFont(font)
                key_item.setBackground(Qt.GlobalColor.lightGray)
                value_item.setBackground(Qt.GlobalColor.lightGray)
            
            self.company_total_table.setItem(row, 0, key_item)
            self.company_total_table.setItem(row, 1, value_item)

        # Tablo başlıklarını ayarla
        header = self.company_total_table.horizontalHeader()
        header.setSectionResizeMode(0, header.ResizeMode.Stretch)
        header.setSectionResizeMode(1, header.ResizeMode.Stretch)

        header = self.employee_details_table.horizontalHeader()
        for i in range(8):
            header.setSectionResizeMode(i, header.ResizeMode.ResizeToContents)







class PDFGenerator:
    def __init__(self, base_dir='docs/employee_files'):  # employee_files -> docs/employee_files
        self.base_dir = base_dir
        self._ensure_base_directory()
        # Helvetica font'unu kullan (ReportLab'in varsayılan fontlarından biri)
        self.font_name = 'Helvetica'

    def _ensure_base_directory(self):
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)

    def _get_turkish_month_name(self, month):
        """Ay numarasına göre Türkçe ay ismi döndür"""
        turkish_months = {
            1: "Ocak", 2: "Subat", 3: "Mart", 4: "Nisan", 5: "Mayis", 6: "Haziran",
            7: "Temmuz", 8: "Agustos", 9: "Eylul", 10: "Ekim", 11: "Kasim", 12: "Aralik"
        }
        return turkish_months.get(month, "")

    def generate_monthly_report(self, employee_data, time_records, summary_data, month, year):
        try:
            employee_folder = f"{employee_data[1]}_{employee_data[2]}_{employee_data[0]}"
            folder_path = os.path.join(self.base_dir, employee_folder)

            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            month_name = self._get_turkish_month_name(month)
            filename = os.path.join(folder_path, f"rapor_{year}_{month}_{month_name}.pdf")

            doc = SimpleDocTemplate(
                filename,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1,
                fontName='Times-Roman'  # Font değiştirildi
            )

            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                spaceBefore=20,
                fontName='Times-Roman'  # Font değiştirildi
            )

            elements = []

            title = Paragraph(f"Aylik Calisma Raporu - {month_name} {year}", title_style)
            elements.append(title)

            elements.append(Paragraph("Calisan Bilgileri", subtitle_style))
            employee_info = [
                ["Ad Soyad:", f"{employee_data[1]} {employee_data[2]}"],
                ["Pozisyon:", employee_data[3]],
                ["Maas:", f"{float(employee_data[4].replace(',', '')):,.0f} TL"],
            ]

                   
            table_style = TableStyle([
                ('FONTNAME', (0,0), (-1,-1), 'Times-Roman'),  # Font değiştirildi
                ('FONTSIZE', (0,0), (-1,-1), 11),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('PADDING', (0,0), (-1,-1), 6),
            ])

            t = Table(employee_info, colWidths=[4*cm, 12*cm])
            t.setStyle(table_style)
            elements.append(t)
            elements.append(Spacer(1, 20))

            elements.append(Paragraph("Aylik Calisma Özeti", subtitle_style))
            summary_rows = [
                ["Toplam Calisan Gun:", f"{summary_data[0]} gun"],
                ["Toplam Calisma Saati:", f"{summary_data[1]:.0f} saat"],
                ["Toplam Fazla Mesai:", f"{summary_data[2]:.0f} saat"],
                ["Normal Mesai Ucreti:", f"{summary_data[3]:,.0f} TL"],
                ["Fazla Mesai Ucreti:", f"{summary_data[4]:,.0f} TL"],
                ["Toplam Ucret:", f"{summary_data[5]:,.0f} TL"]
            ]

            summary_table = Table(summary_rows, colWidths=[6*cm, 10*cm])
            summary_table.setStyle(table_style)
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            elements.append(Paragraph("Detayli Mesai Kayitlari", subtitle_style))
            headers = ['Tarih', 'Giris', 'Cikis', 'Sure', 'Mesai', 'Tip', 'Ucret']
            detail_data = [headers]

            for record in time_records:
                # Önce mesai tipini belirle
                if record[14]:  # is_weekend_leave
                    mesai_tipi = "Haftalik Izin"  # Türkçe karakter düzeltildi
                    mesai_durumu = "Haftalik Izin"  # Türkçe karakter düzeltildi
                elif record[7]:  # is_holiday
                    mesai_tipi = "Tatil"
                    mesai_durumu = "Tatil"
                elif record[12]:  # is_sick_leave
                    mesai_tipi = "Raporlu"
                    mesai_durumu = "Raporlu"
                elif record[13]:  # is_paid_leave
                    mesai_tipi = "Izinli"  # Türkçe karakter düzeltildi
                    mesai_durumu = "Izinli"  # Türkçe karakter düzeltildi
                else:
                    if record[8]:  # is_weekend
                        mesai_tipi = "Cumartesi"
                    else:
                        mesai_tipi = "Normal"

                    standard_hours = 5 if record[8] else 8
                    if record[5] < standard_hours:
                        mesai_durumu = f"Eksik ({standard_hours - record[5]:.1f}s)"
                    else:
                        mesai_durumu = f"Fazla ({record[6]:.1f}s)" if record[6] > 0 else "Tam"

                row = [
                    record[2],
                    record[3],
                    record[4],
                    f"{record[5]:.1f}s",
                    mesai_durumu,
                    mesai_tipi,
                    f"{record[11]:,.1f} TL"
                ]
                detail_data.append(row)






            
            detail_table = Table(detail_data, colWidths=[3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm, 2.5*cm, 3*cm])
            detail_style = TableStyle([
                ('FONTNAME', (0,0), (-1,-1), self.font_name),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('PADDING', (0,0), (-1,-1), 4),
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ])
            detail_table.setStyle(detail_style)
            elements.append(detail_table)
            
            doc.build(elements)
            return filename
            
        except Exception as e:
            print(f"PDF olusturma hatasi: {str(e)}")
            raise e



class ActivityLogTab(QWidget):
    def __init__(self, db, main_window=None):
        super().__init__()
        self.db = db
        self.main_window = main_window
        self.file_system = EmployeeFileSystem(db=db, base_dir='docs/employee_files')
        self.init_ui()


    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)
    
        # Üst kontrol paneli
        control_panel = QWidget()
        control_layout = QHBoxLayout()
    
        # Tarih seçici
        self.date_select = QDateEdit()
        self.date_select.setCalendarPopup(True)
        self.date_select.setDate(QDate.currentDate())
        self.date_select.dateChanged.connect(self.load_logs)
    
        # Yenile butonu
        refresh_button = QPushButton("Yenile")
        refresh_button.clicked.connect(self.load_logs)
        refresh_button.setIconSize(QSize(20, 20))  # Set icon size to 20x20 pixels
        refresh_button.setIcon(QIcon("docs/icon/update.png"))
    
        control_layout.addWidget(QLabel("Tarih:"))
        control_layout.addWidget(self.date_select)
        control_layout.addWidget(refresh_button)
        control_layout.addStretch()
        control_panel.setLayout(control_layout)
    
        # Log tablosu
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(6)
        self.log_table.setHorizontalHeaderLabels([
            'Tarih/Saat', 
            'Kullanıcı', 
            'İşlem Tipi', 
            'Değişiklik Detayları', 
            'İlgili Bölüm',
            'Kayıt ID'
        ])
        setup_modern_table(self.log_table)
        self.log_table.itemDoubleClicked.connect(self.show_correction_dialog)
    
        # Arama alanı
        search_container = QWidget()
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Arama yap...")
        self.search_input.textChanged.connect(self.filter_logs)
        search_layout.addWidget(QLabel("Ara:"))
        search_layout.addWidget(self.search_input)
        search_container.setLayout(search_layout)
    
        # Layout'a widget'ları ekle
        layout.addWidget(control_panel)
        layout.addWidget(search_container)
        layout.addWidget(self.log_table)
        self.setLayout(layout)
    
        # Başlangıç yüklemesi
        self.load_logs()

    



    # ActivityLogTab sınıfının show_correction_dialog metodu
    def show_correction_dialog(self, item):
        row = item.row()
        action_type = self.log_table.item(row, 2).text()
        record_id = self.log_table.item(row, 5).text()
        related_table = self.log_table.item(row, 4).text()
        description = self.log_table.item(row, 3).text()

        # Mesai kayıtlarıyla ilgili işlemler için düzenleme imkanı sağla
        if "Mesai" in related_table and ("Mesai Kaydı Ekleme" in action_type or "Mesai Kaydı Güncelleme" in action_type):
            dialog = QDialog(self)
            try:
                cursor = self.db.conn.cursor()
                cursor.execute('''
                    SELECT t.*, e.name, e.surname
                    FROM time_records t
                    JOIN employees e ON t.employee_id = e.id
                    WHERE t.id = ?
                ''', (record_id,))
                record = cursor.fetchone()

                if record:
                    dialog = QDialog(self)
                    dialog.setWindowTitle('Mesai Kaydı Detayları')
                    layout = QVBoxLayout()

                    # Form container
                    form_container = QWidget()
                    form_layout = QFormLayout()

                    # Çalışan bilgileri 
                    info_label = QLabel(f"Çalışan: {record[-2]} {record[-1]}\nTarih: {record[2]}")
                    form_layout.addRow(info_label)

                    # Giriş ve çıkış saatleri için input alanları
                    self.check_in_input = QLineEdit(record[3])
                    self.check_out_input = QLineEdit(record[4])
                    form_layout.addRow("Giriş Saati:", self.check_in_input)
                    form_layout.addRow("Çıkış Saati:", self.check_out_input)

                    # Checkbox'lar
                    self.holiday_checkbox = QCheckBox("Tatil/Bayram Günü")
                    self.sick_leave_checkbox = QCheckBox("Raporlu")
                    self.paid_leave_checkbox = QCheckBox("İzinli")
                    self.weekend_leave_checkbox = QCheckBox("Haftalık İzin")

                    # Mevcut duruma göre checkbox'ları ayarla
                    self.holiday_checkbox.setChecked(record[7])  # is_holiday
                    self.sick_leave_checkbox.setChecked(record[12])  # is_sick_leave
                    self.paid_leave_checkbox.setChecked(record[13])  # is_paid_leave
                    self.weekend_leave_checkbox.setChecked(record[14])  # is_weekend_leave

                    # Checkbox'ları form'a ekle
                    form_layout.addRow(self.holiday_checkbox)
                    form_layout.addRow(self.sick_leave_checkbox)
                    form_layout.addRow(self.paid_leave_checkbox)
                    form_layout.addRow(self.weekend_leave_checkbox)

                    # Değişiklik nedeni input'u
                    self.reason_input = QLineEdit()
                    self.reason_input.setPlaceholderText("Düzeltme nedenini yazın...")
                    form_layout.addRow("Düzeltme Nedeni:", self.reason_input)

                    # Değişiklik detaylarını gösterecek label
                    record_history_label = QLabel()
                    record_history_label.setWordWrap(True)
                    record_history_label.setStyleSheet("""
                        QLabel {
                            color: #333333;
                            font-size: 12px;
                            padding: 15px;
                            background-color: #f5f5f5;
                            border-radius: 5px;
                            border: 1px solid #ddd;
                            min-height: 100px;
                        }
                    """)

                    cursor.execute('''
                        SELECT action_description, timestamp 
                        FROM activity_logs 
                        WHERE related_id = ? 
                        AND (action_type = 'Mesai Kaydı Güncelleme' OR action_type = 'Mesai Kaydı Ekleme')
                        ORDER BY timestamp DESC
                    ''', (record_id,))
                    
                    changes = cursor.fetchall()
                    if changes:
                        change_history = "DEĞİŞİKLİK GEÇMİŞİ\n" + "="*50 + "\n\n"
                        for idx, (change_desc, timestamp) in enumerate(changes, 1):
                            # Timestamp'i düzenli formata çevir
                            dt = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                            formatted_date = dt.strftime('%d.%m.%Y %H:%M')
                            
                            # Her değişiklik kaydındaki saat bilgilerini formatla
                            formatted_change = self.format_log_content(change_desc)
                            
                            # Değişiklik başlığını oluştur
                            change_history += f"Değişiklik #{idx} - {formatted_date}\n"
                            change_history += "-"*50 + "\n"
                            
                            # İçeriği düzenli bir şekilde ekle
                            lines = formatted_change.split('\n')
                            for line in lines:
                                if any(header in line for header in ['Çalışan:', 'Tarih:', 'Yapılan Değişiklikler:', 
                                                                   'Önceki Durum:', 'Yeni Durum:', 'Düzeltme Nedeni:']):
                                    # Başlıkları kalın yap ve ayır
                                    change_history += f"\n• {line.strip()}\n"
                                else:
                                    # Normal satırları girintili ekle
                                    change_history += f"    {line.strip()}\n"
                            
                            change_history += "\n" + "="*50 + "\n\n"
                        
                        # Önceki label'ı QTextEdit ile değiştir
                        history_text = QTextEdit()
                        history_text.setReadOnly(True)
                        history_text.setMinimumHeight(300)
                        history_text.setStyleSheet("""
                            QTextEdit {
                                font-family: Arial;
                                font-size: 12px;
                                line-height: 1.5;
                                padding: 10px;
                                background-color: #f8f9fa;
                                border: 1px solid #e9ecef;
                                border-radius: 5px;
                            }
                        """)
                        history_text.setText(change_history)
                        
                        # QTextEdit'i layout'a ekle
                        form_layout.addRow(history_text)


                    # Kaydet ve İptal butonları
                    button_layout = QHBoxLayout()
                    save_button = QPushButton("Kaydet")
                    save_button.clicked.connect(lambda: self.save_correction(
                        dialog, record_id, record[1], record[2]  # record[1] is employee_id, record[2] is date
                    ))
                    cancel_button = QPushButton("İptal")
                    cancel_button.clicked.connect(dialog.reject)

                    button_layout.addWidget(save_button)
                    button_layout.addWidget(cancel_button)
                    form_layout.addRow(button_layout)

                    # Form container'ı ana layout'a ekle
                    form_container.setLayout(form_layout)
                    layout.addWidget(form_container)

                    dialog.setLayout(layout)
                    dialog.setMinimumWidth(600)
                    dialog.setMinimumHeight(800)
                    dialog.exec()

            except Exception as e:
                QMessageBox.warning(self, 'Hata', f'Kayıt yüklenirken hata oluştu: {str(e)}')
        else:
            # Diğer kayıtlar için sadece görüntüleme dialogu
            dialog = QDialog(self)
            dialog.setWindowTitle(f'{related_table} - Değişiklik Detayları')
            dialog.setMinimumWidth(500)
            dialog.setMinimumHeight(400)
            layout = QVBoxLayout()

            # Başlık
            title_label = QLabel(action_type)
            title_label.setStyleSheet("""
                QLabel {
                    font-size: 14px;
                    font-weight: bold;
                    color: #333333;
                    padding: 10px;
                    background-color: #f8f9fa;
                    border-radius: 5px;
                    margin-bottom: 10px;
                }
            """)
            layout.addWidget(title_label)

            # Değişiklik detayları
            detail_text = QTextEdit()
            detail_text.setReadOnly(True)
            detail_text.setStyleSheet("""
                QTextEdit {
                    font-size: 12px;
                    padding: 15px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    background-color: white;
                    line-height: 1.5;
                }
            """)
            detail_text.setText(description)
            layout.addWidget(detail_text)

            # Kapat butonu
            close_button = QPushButton("Kapat")
            close_button.clicked.connect(dialog.accept)
            close_button.setStyleSheet("""
                QPushButton {
                    background-color: #CC3300;
                    color: white;
                    padding: 8px 16px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    min-width: 100px;
                    margin-top: 10px;
                }
                QPushButton:hover {
                    background-color: #E63900;
                }
            """)
            layout.addWidget(close_button, alignment=Qt.AlignmentFlag.AlignCenter)

            dialog.setLayout(layout)
            dialog.exec()


    def get_previous_state(self, record_id):
        """Mesai kaydının önceki durumunu al"""
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT 
                t.check_in, t.check_out, t.working_hours, t.overtime_hours,
                t.is_holiday, t.is_sick_leave, t.is_paid_leave, t.is_weekend_leave,
                e.name, e.surname
            FROM time_records t
            JOIN employees e ON t.employee_id = e.id
            WHERE t.id = ?
        ''', (record_id,))
        return cursor.fetchone()
    
    def format_state_description(self, state):
        """Durum bilgisini formatla"""
        status = []
        if state[4]:  # is_holiday
            status.append("Tatil/Bayram")
        if state[5]:  # is_sick_leave
            status.append("Raporlu")
        if state[6]:  # is_paid_leave
            status.append("İzinli")
        if state[7]:  # is_weekend_leave
            status.append("Haftalık İzin")
        
        return {
            'check_in': state[0],
            'check_out': state[1],
            'working_hours': f"{state[2]:.1f} saat",
            'overtime_hours': f"{state[3]:.1f} saat",
            'status': ' + '.join(status) if status else "Normal Mesai",
            'employee': f"{state[8]} {state[9]}"
        }



    def get_record_history(self, record_id):
        """Bir mesai kaydının düzenleme geçmişini alır"""
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT action_description, timestamp
            FROM activity_logs 
            WHERE related_id = ? AND action_type = 'Mesai Kaydı Güncelleme'
            ORDER BY timestamp DESC
            LIMIT 1
        ''', (record_id,))
        return cursor.fetchone()

    



    def save_correction(self, dialog, record_id, employee_id, date):
        """Mesai düzeltmelerini kaydet ve logla"""
        try:
            # Düzeltme nedeni kontrolü
            if not self.reason_input.text():
                QMessageBox.warning(dialog, 'Hata', 'Düzeltme nedeni gereklidir!')
                return

            # Yeni değerleri al
            check_in = self.check_in_input.text()
            check_out = self.check_out_input.text()
            is_holiday = self.holiday_checkbox.isChecked()
            is_sick_leave = self.sick_leave_checkbox.isChecked()
            is_paid_leave = self.paid_leave_checkbox.isChecked()
            is_weekend_leave = self.weekend_leave_checkbox.isChecked()

            # Önceki kaydı al
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT t.*, e.name, e.surname
                FROM time_records t
                JOIN employees e ON t.employee_id = e.id
                WHERE t.id = ?
            ''', (record_id,))
            old_record = cursor.fetchone()

            if not old_record:
                QMessageBox.warning(dialog, 'Hata', 'Kayıt bulunamadı!')
                return

            # Değişiklikleri belirle
            changes = []
            if old_record[3] != check_in:
                changes.append(f"Giriş Saati: {old_record[3]} -> {check_in}")
            if old_record[4] != check_out:
                changes.append(f"Çıkış Saati: {old_record[4]} -> {check_out}")
            if old_record[7] != is_holiday:
                changes.append(f"Tatil/Bayram: {'Evet' if old_record[7] else 'Hayır'} -> {'Evet' if is_holiday else 'Hayır'}")
            if old_record[12] != is_sick_leave:
                changes.append(f"Raporlu: {'Evet' if old_record[12] else 'Hayır'} -> {'Evet' if is_sick_leave else 'Hayır'}")
            if old_record[13] != is_paid_leave:
                changes.append(f"İzinli: {'Evet' if old_record[13] else 'Hayır'} -> {'Evet' if is_paid_leave else 'Hayır'}")
            if old_record[14] != is_weekend_leave:
                changes.append(f"Haftalık İzin: {'Evet' if old_record[14] else 'Hayır'} -> {'Evet' if is_weekend_leave else 'Hayır'}")

            if not changes:
                QMessageBox.warning(dialog, 'Uyarı', 'Hiçbir değişiklik yapılmadı!')
                return

            # Çalışma saatlerini hesapla
            working_hours, overtime_hours = self.db.calculate_working_hours(check_in, check_out, date)
            is_weekend = self.db.is_weekend(date)

            # Ücret hesaplamaları
            if is_paid_leave or is_weekend_leave:
                working_hours = 0
                overtime_hours = 0
                base_pay = 0
                overtime_pay = 0

                if is_weekend_leave:
                    cursor.execute('SELECT salary FROM employees WHERE id = ?', (employee_id,))
                    monthly_salary = cursor.fetchone()[0]
                    monthly_salary = float(monthly_salary.replace(',', ''))
                    total_pay = monthly_salary / 30
                else:
                    total_pay = 0
            elif is_sick_leave:
                working_hours = 8
                overtime_hours = 0
                base_pay = 0
                overtime_pay = 0
                total_pay = 0
            else:
                base_pay, overtime_pay, total_pay = self.db.calculate_pay(
                    employee_id, working_hours, overtime_hours, is_holiday, is_weekend
                )

            # Veritabanını güncelle
            cursor.execute('''
                UPDATE time_records 
                SET check_in = ?, check_out = ?, working_hours = ?, 
                    overtime_hours = ?, is_holiday = ?, base_pay = ?, 
                    overtime_pay = ?, total_pay = ?, is_sick_leave = ?,
                    is_paid_leave = ?, is_weekend_leave = ?
                WHERE id = ?
            ''', (
                check_in, check_out, working_hours, overtime_hours, 
                is_holiday, base_pay, overtime_pay, total_pay, 
                is_sick_leave, is_paid_leave, is_weekend_leave, record_id
            ))

            # Excel dosyasını güncelle
            try:
                record_data = {
                    'date': date,
                    'check_in': check_in,
                    'check_out': check_out,
                    'working_hours': working_hours,
                    'regular_hours': working_hours - overtime_hours,
                    'overtime_hours': overtime_hours,
                    'missing_hours': max(0, 8 - working_hours),
                    'work_type': 'Haftalık İzin' if is_weekend_leave else (
                        'Raporlu' if is_sick_leave else (
                        'İzinli' if is_paid_leave else (
                        'Tatil' if is_holiday else 'Normal'))),
                    'regular_pay': base_pay,
                    'overtime_pay': overtime_pay,
                    'missing_pay': 0,
                    'total_pay': total_pay
                }

                self.file_system.update_employee_timesheet(
                    employee_id,
                    old_record[-2],  # name
                    old_record[-1],  # surname
                    record_data
                )

            except Exception as excel_error:
                print(f"Excel güncelleme hatası: {str(excel_error)}")
                QMessageBox.warning(
                    dialog, 
                    'Excel Hatası',
                    'Mesai kaydı güncellendi ancak Excel dosyası güncellenirken hata oluştu!'
                )

            # Log kaydı ekle
            if self.main_window and self.main_window.current_username:
                status_comparison = (
                    f"Önceki Durum:\n"
                    f"Çalışma Süresi: {format_time_as_hours_minutes(old_record[5])}\n"
                    f"Fazla Mesai: {format_time_as_hours_minutes(old_record[6])}\n"
                    f"Ücret: {old_record[11]:.2f} TL\n\n"
                    f"Yeni Durum:\n"
                    f"Çalışma Süresi: {format_time_as_hours_minutes(working_hours)}\n"
                    f"Fazla Mesai: {format_time_as_hours_minutes(overtime_hours)}\n"
                    f"Ücret: {total_pay:.2f} TL"
                )

                changes_description = (
                    f"Çalışan: {old_record[-2]} {old_record[-1]}\n"
                    f"Tarih: {date}\n\n"
                    f"Yapılan Değişiklikler:\n"
                    f"{chr(10).join(changes)}\n\n"
                    f"{status_comparison}\n\n"
                    f"Düzeltme Nedeni: {self.reason_input.text()}"
                )

                self.db.log_activity(
                    self.main_window.current_username,
                    "TIME_RECORD_UPDATE",
                    changes_description,
                    "time_records",
                    record_id
                )

            self.db.conn.commit()
            dialog.accept()
            self.load_logs()
            QMessageBox.information(self, 'Başarılı', 'Mesai kaydı düzeltildi!')

        except ValueError:
            QMessageBox.warning(self, 'Hata', 'Geçersiz saat formatı! Lütfen HH:MM formatında girin.')
        except Exception as e:
            print(f"Hata: {str(e)}")
            QMessageBox.warning(self, 'Hata', f'Düzeltme hatası: {str(e)}')




    def get_status_text(self, record):
        """Durum metnini oluştur"""
        if isinstance(record, dict):
            if record['is_holiday']:
                return "Tatil/Bayram"
            elif record['is_sick_leave']:
                return "Raporlu"
            elif record['is_paid_leave']:
                return "İzinli"
            elif record['is_weekend_leave']:
                return "Haftalık İzin"
            else:
                return "Normal Mesai"
        else:
            if record[7]:  # is_holiday
                return "Tatil/Bayram"
            elif record[12]:  # is_sick_leave
                return "Raporlu"
            elif record[13]:  # is_paid_leave
                return "İzinli"
            elif record[14]:  # is_weekend_leave
                return "Haftalık İzin"
            else:
                return "Normal Mesai"

    def get_user_id(self):
        """Mevcut kullanıcının ID'sini al"""
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id FROM users WHERE username = ?', 
                      (self.main_window.current_username,))
        result = cursor.fetchone()
        return result[0] if result else None




    def load_logs(self):
        selected_date = self.date_select.date().toPyDate()
        cursor = self.db.conn.cursor()

        cursor.execute('''
            SELECT 
                datetime(timestamp, 'localtime'),
                username,
                action_type,
                action_description,
                related_table,
                related_id
            FROM activity_logs
            WHERE date(timestamp) = ?
            ORDER BY timestamp DESC
        ''', (selected_date.strftime('%Y-%m-%d'),))

        logs = cursor.fetchall()
        self.log_table.setRowCount(len(logs))

        for row, log in enumerate(logs):
            for col, value in enumerate(log):
                # Eğer bu alan action_description ise sadeleştirilmiş açıklama göster
                if col == 3:
                    if "Mesai Kaydı Güncelleme" in log[2]:
                        # Yapılan değişiklikleri bul
                        changes = []
                        if "Giriş Saati:" in str(value):
                            changes.append("Giriş saati değiştirildi")
                        if "Çıkış Saati:" in str(value):
                            changes.append("Çıkış saati değiştirildi")
                        if "Tatil/Bayram:" in str(value):
                            changes.append("Tatil durumu değiştirildi")
                        if "Raporlu:" in str(value):
                            changes.append("Rapor durumu değiştirildi")
                        if "İzinli:" in str(value):
                            changes.append("İzin durumu değiştirildi")

                        # Değişiklikleri kısa bir metin olarak göster
                        summary = " | ".join(changes) if changes else "Güncelleme yapıldı"
                        item = QTableWidgetItem(summary)
                    else:
                        item = QTableWidgetItem(str(value))
                else:
                    item = QTableWidgetItem(str(value) if value is not None else '')

                self.log_table.setItem(row, col, item)

        # Sütun genişliklerini ayarla
        header = self.log_table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        for i in range(self.log_table.columnCount()):
            if i != 3:
                header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)


    def format_log_content(self, content):
        """Log içeriğini daha minimalist formatta düzenle"""
        if not content:
            return ""
    
        # Temel bilgileri ayıkla
        lines = content.split('\n')
        formatted_lines = []
        current_section = ""
    
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Çalışma süresi formatını düzenle
            if "Çalışma Süresi:" in line:
                parts = line.split(":")
                if len(parts) == 2:
                    try:
                        hours = float(parts[1].replace("saat", "").strip())
                        formatted_time = format_time_as_hours_minutes(hours)
                        line = f"Çalışma Süresi: {formatted_time}"
                    except ValueError:
                        pass
            # Fazla mesai formatını düzenle
            elif "Fazla Mesai:" in line:
                parts = line.split(":")
                if len(parts) == 2:
                    try:
                        hours = float(parts[1].replace("saat", "").strip())
                        formatted_time = format_time_as_hours_minutes(hours)
                        line = f"Fazla Mesai: {formatted_time}"
                    except ValueError:
                        pass
            
            # Başlıkları kontrol et
            if "Çalışan:" in line:
                formatted_lines.append(f"👤 {line}")
            elif "Tarih:" in line:
                formatted_lines.append(f"📅 {line}")
            elif "Yapılan Değişiklikler:" in line:
                current_section = "changes"
                formatted_lines.append("\n📝 DEĞİŞİKLİKLER")
            elif "Önceki Durum:" in line:
                current_section = "before"
                formatted_lines.append("\n⏮️ ÖNCEKİ DURUM")
            elif "Yeni Durum:" in line:
                current_section = "after"
                formatted_lines.append("\n⏭️ YENİ DURUM")
            elif "Düzeltme Nedeni:" in line:
                current_section = "reason"
                formatted_lines.append(f"\n❓ {line}")
            elif " ->" in line:
                # Çalışma süresi değişikliklerini formatla
                if "saat ->" in line:
                    parts = line.split("->")
                    try:
                        before = float(parts[0].split("saat")[0].strip())
                        after = float(parts[1].strip().split("saat")[0].strip())
                        formatted_before = format_time_as_hours_minutes(before)
                        formatted_after = format_time_as_hours_minutes(after)
                        line = f"{parts[0].split(':')[0]}: {formatted_before} -> {formatted_after}"
                    except ValueError:
                        pass
                formatted_lines.append(f"  ➜ {line}")
            else:
                # Diğer satırları girintili göster
                formatted_lines.append(f"    {line}")
    
        return "\n".join(formatted_lines)

    def filter_logs(self):
        search_text = self.search_input.text().lower()
        for row in range(self.log_table.rowCount()):
            row_visible = False
            for col in range(self.log_table.columnCount()):
                item = self.log_table.item(row, col)
                if item and search_text in item.text().lower():
                    row_visible = True
                    break
            self.log_table.setRowHidden(row, not row_visible)

    def refresh_logs(self):
        self.load_logs()

def format_time_as_hours_minutes(decimal_hours):
    """
    Ondalıklı saat değerini saat ve dakika formatına dönüştürür.
    Örnek: 1.75 -> "1 saat 45 dakika"
    """
    hours = int(decimal_hours)
    minutes = int((decimal_hours - hours) * 60)
    
    if hours == 0:
        if minutes == 0:
            return "0 saat"
        return f"{minutes} dakika"
    elif minutes == 0:
        return f"{hours} saat"
    else:
        return f"{hours} saat {minutes} dakika"

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.current_username = None
        self.user_permissions = None
        self.show_login()
        self.setWindowIcon(QIcon("docs/icon/irem.jpg"))

    def init_ui(self):
        self.setWindowTitle('İremin programı')
        self.setGeometry(180, 60, 1200, 750)
    
        # StyleEditor'ı başlat ama gösterme
        self.style_editor = StyleEditor()
        
        # Ana stil tanımlamasını uygula
        self.setStyleSheet(MAIN_STYLE)
    
        # Tab widget oluştur
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)
    
        # Departmana göre sekmeleri ekle
        department = self.user_permissions.get('department')
    
        # Sekmeleri oluştur ve bağlantıları kur
        if department == 'Admin':
            # Admin tüm sekmeleri görebilir
            self.employee_tab = EmployeeTab(self.db, self)
            self.time_record_tab = TimeRecordTab(self.db, self)
            self.time_record_tab.main_window = self
            self.report_tab = ReportingTab(self.db)
            self.company_report_tab = CompanyReportTab(self.db)
            self.user_management_tab = UserManagementTab(self.db, self)
            self.activity_log_tab = ActivityLogTab(self.db)
    
            self.tab_widget.setIconSize(QSize(20, 20))
    
            employee_icon = QIcon("docs/icon/employees.png")
            time_icon = QIcon("docs/icon/timeeadd.png")
            report_icon = QIcon("docs/icon/user-analytics.png")
            company_icon = QIcon("docs/icon/report.png")
            user_icon = QIcon("docs/icon/process.png")
            log_icon = QIcon("docs/icon/transactionhistory.png")
    
            # Sekmeleri ekle
            self.tab_widget.addTab(self.employee_tab, employee_icon, "Çalışanlar")
            self.tab_widget.addTab(self.time_record_tab, time_icon, "Mesai Kayıtları")
            self.tab_widget.addTab(self.report_tab, report_icon, "Çalışan Raporları")
            self.tab_widget.addTab(self.company_report_tab, company_icon, "İşletme Raporları")
            self.tab_widget.addTab(self.user_management_tab, user_icon, "Kullanıcı Yönetimi")
            self.activity_log_tab = ActivityLogTab(self.db, self)
            self.tab_widget.addTab(self.activity_log_tab, log_icon, "İşlem Geçmişi")
    
            # Admin için tüm sinyalleri bağla
            self.employee_tab.employee_added.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_added.connect(self.report_tab.update_employee_select)
            self.employee_tab.employee_updated.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_updated.connect(self.report_tab.update_employee_select)
            self.employee_tab.employee_deleted.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_deleted.connect(self.report_tab.update_employee_select)
            
            # Mesai kayıtları sinyallerini bağla
            self.time_record_tab.time_record_added.connect(lambda: self.activity_log_tab.refresh_logs())
            self.time_record_tab.time_record_updated.connect(lambda: self.activity_log_tab.refresh_logs())
            self.time_record_tab.time_record_deleted.connect(lambda: self.activity_log_tab.refresh_logs())
            self.time_record_tab.time_record_added.connect(self.activity_log_tab.load_logs)
            self.time_record_tab.time_record_updated.connect(self.activity_log_tab.load_logs)
            
            # Rapor sinyallerini bağla
            self.report_tab.report_generated.connect(lambda: self.activity_log_tab.refresh_logs())
            self.report_tab.report_updated.connect(lambda: self.activity_log_tab.refresh_logs())
    
        elif department == 'Yönetici':
            # Yönetici sadece ilgili sekmeleri görebilir
            self.employee_tab = EmployeeTab(self.db)
            self.time_record_tab = TimeRecordTab(self.db)
            self.time_record_tab.main_window = self
            self.report_tab = ReportingTab(self.db)
            self.company_report_tab = CompanyReportTab(self.db)
            
            # Sekmeleri ekle
            self.tab_widget.addTab(self.employee_tab, "Çalışanlar")
            self.tab_widget.addTab(self.time_record_tab, "Mesai Kayıtları")
            self.tab_widget.addTab(self.report_tab, "Çalışan Raporları")
            self.tab_widget.addTab(self.company_report_tab, "İşletme Raporları")
            
            # Yönetici için sinyalleri bağla
            self.employee_tab.employee_added.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_added.connect(self.report_tab.update_employee_select)
            self.employee_tab.employee_updated.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_updated.connect(self.report_tab.update_employee_select)
            self.employee_tab.employee_deleted.connect(self.time_record_tab.update_employee_select)
            self.employee_tab.employee_deleted.connect(self.report_tab.update_employee_select)
            
            # Mesai kayıtları sinyallerini bağla
            self.time_record_tab.time_record_added.connect(self.report_tab.load_report)
            self.time_record_tab.time_record_updated.connect(self.report_tab.load_report)
            self.time_record_tab.time_record_deleted.connect(self.report_tab.load_report)
            
        elif department == 'Muhasebe':
            # Muhasebe sadece mesai kayıtlarını görebilir
            self.time_record_tab = TimeRecordTab(self.db)
            self.time_record_tab.main_window = self
            self.tab_widget.addTab(self.time_record_tab, "Mesai Kayıtları")

        # Pencereyi göster
        self.show()

    def show_login(self):
        while True:  # Sürekli login dialogu göster
            login_dialog = LoginDialog(self.db)
            result = login_dialog.exec()
            
            if result == QDialog.DialogCode.Accepted:  # Başarılı giriş
                self.user_permissions = login_dialog.user_permissions
                self.current_username = login_dialog.username
                self.init_ui()
                break  # Başarılı girişte döngüden çık
            else:  # Başarısız giriş
                reply = QMessageBox.question(
                    self,
                    'Yardım İrem',
                    'Yanlış şifre lütfen İremden şifrenizi alınız',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    sys.exit()

    def restart_application(self):
        """Uygulamayı yeniden başlat"""
        QApplication.quit()
        os.execl(sys.executable, sys.executable, *sys.argv)

    def closeEvent(self, event):
        self.db.close()
        event.accept()
def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()